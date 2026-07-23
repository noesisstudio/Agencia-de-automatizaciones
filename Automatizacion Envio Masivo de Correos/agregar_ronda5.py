"""Añade a cola_envios.xlsx la ronda 5 (Olot, Blanes, Tortosa — búsqueda 17/07/2026),
deduplicando contra TODOS los emails ya presentes en la cola y en los archivos de
las tandas anteriores. Las que tienen email quedan con fecha_prevista de hoy
(2026-07-17) para completar el presupuesto del día; las 'sin email' se guardan
como registro. Después ejecutar: reconstruir_maestro.py"""
from openpyxl import load_workbook

FECHA_HOY = "2026-07-17"

# (empresa, ciudad, email_o_None, fuente)
RONDA5 = [
    ("Aulinas Assessors", "Olot", "gemma@aulinas.cat", "https://aulinas.cat/es/contacto/"),
    ("Gestoria Castells", "Olot", "administracio@gestoriacastells.cat", "https://www.gestoriacastells.cat/"),
    ("Agència CEC", "Olot", None, "https://www.paginasamarillas.es/f/olot/agencia-cec_009531096_000000001.html"),
    ("Capdevila Advocats i Economistes", "Olot", "info@capdevilaeconomistes.com", "https://www.holded.com/es/asesorias/capdevila-economistes-slp"),
    ("Assessoria Oliveras SL", "Olot", "assessoria@assol.com", "https://assol.com/contacte/"),
    ("J&S Serveis i Gestions", "Olot", "jsserveisigestions@gmail.com", "http://serveisigestionsjs.blogspot.com/p/contacta.html"),
    ("Jose Pauli Costa", "Olot", None, "https://www.paginasamarillas.es/f/olot/jose-pauli-costa_009496670_000000002.html"),
    ("Gestoria Girgas SL", "Olot", "info@gestoriagirgas.com", "https://gestoriagirgas.com/"),
    ("Forum Assessors", "Olot", "info@forumsl.com", "https://www.infoasesorias.es/forum-assessors/"),
    ("Assessors Garrotxa", "Olot", "info@assessorsgarrotxa.com", "https://www.assessorsgarrotxa.com/es/contacto/"),
    ("Ribas Àlvarez Assessors (oficina Olot)", "Olot", "assessoria@ribasalvarez.cat", "https://www.holded.com/es/asesorias/ribas-alvarez-assessors-i-consultors-sl"),
    ("Margarita Pascual Agusti", "Olot", None, "https://www.paginasamarillas.es/f/olot/margarita-pascual-agusti_207724493_000000001.html"),
    ("Clos Assessors SL", "Olot", None, "https://empresite.eleconomista.es/CLOS-ASSESSORS.html"),
    ("Garrotxa Activa", "Olot", "info@garrotxaactiva.cat", "https://www.garrotxaactiva.cat/contacte-i-situacio-assessoria-olot/"),
    ("Assessoria Gabinet Permar SL", "Olot", None, "https://www.paginasamarillas.es/search/asesoria-laboral/all-ma/girona/all-is/olot"),
    ("Assessoria Tax Olot", "Olot", "central@tax.es", "https://es.cybo.com/ES-biz/assessoria-tax-olot"),
    ("BM Gestoría", "Olot", "bmuro@bmgestoria.com", "https://bmgestoria.com/"),
    ("Finexen", "Olot", None, "https://www.gestorias.es/girona/olot"),
    ("Adade Assessors", "Olot", None, "https://www.paginasamarillas.es/f/olot/adade-assessors_195404884_000000001.html"),
    ("Agencia O.T.A", "Olot", None, "https://www.gestorias.es/girona/olot/agencia-o-t-a-194"),
    ("Gestió Plural", "Olot", None, "https://www.gestorias.es/girona/olot/gestio-plural-6012"),
    ("Prime Consulting", "Olot", "info@primeconsulting.cat", "https://primeconsulting.cat/contacte/"),

    ("Gestió 2002", "Blanes", "ges@gestio2002.net", "https://www.gestio2002.net/"),
    ("100x100 Pymes", "Blanes", "joan@100x100pymes.com", "https://100x100pymes.com/cas/contactar.html"),
    ("Gabinet Camero & Castiblanque SL", "Blanes", "jce@economistes.com", "https://www.gabinetcamerocastiblanquesl.es/contacto"),
    ("Albano Associats", "Blanes", "info@albanoassociats.com", "https://albanoassociats.com/"),
    ("Assessoria Tax Blanes", "Blanes", None, "https://www.gestorias.es/girona/blanes"),
    ("SA Consulting (Blanes)", "Blanes", None, "https://www.gestorias.es/girona/blanes"),
    ("Buxó Assessors SL", "Blanes", "info@buxo.cat", "https://www.buxo.cat/contacto/"),
    ("Gestarea Consulting", "Blanes", None, "https://www.gestorias.es/girona/blanes"),
    ("AO Assessors", "Blanes", None, "https://www.paginasamarillas.es/f/blanes/ao-assessors_221036999_000000001.html"),
    ("Gestorías Ramos", "Blanes", None, "https://www.gestorias.es/girona/blanes/gestorias-ramos-16"),
    ("MGF Gestions", "Blanes", "mgfgestions@mgfgestions.com", "https://gestoriaenblanes.com/contacto/"),
    ("Som Gestió", "Blanes", None, "https://www.gestorias.es/girona/blanes/som-gestio-2955"),
    ("Gestoría Provença", "Blanes", None, "https://www.gestorias.es/girona/blanes/gestoria-provenca-4335"),
    ("Barcons Assessors SL", "Blanes", None, "https://www.gestorias.es/girona/blanes/barcons-assessors-6216"),
    ("Carles i Prats SL", "Blanes", "carlesipratx@stl.logiccontrol.es", "https://onebusiness.place/es/datos/asesor-de-empresa-blanes-carles-i-prats-s-l-023740160014057031.html"),
    ("GrupRC", "Blanes", "info@gruprc.cat", "https://gruprc.com/contacto/"),
    ("Jiménez & Liébana Assessors", "Blanes", None, "https://www.paginasamarillas.es/search/asesores-fiscales/all-ma/girona/all-is/blanes"),
    ("Miguel Estela", "Blanes", None, "https://www.paginasamarillas.es/f/blanes/miguel-estela_218701563_000000001.html"),
    ("Galván Assessoria", "Blanes", None, "https://www.paginasamarillas.es/f/blanes/galvan-assessoria_006710263_000000001.html"),
    ("Asesoría Chaves", "Blanes", None, "https://www.paginasamarillas.es/f/blanes/asesoria-chaves_020969119_000000001.html"),

    ("Gestoría y Asesoría Mauri", "Tortosa", "mauri@gestoriamauri.com", "https://www.cylex.es/tortosa/gestor%C3%ADa-y-asesor%C3%ADa-mauri-12864391.html"),
    ("Gestoria Lluís Poy", "Tortosa", "carmen@lluispoy.com", "https://www.lluispoy.es/"),
    ("Mulet Lluis SL", "Tortosa", None, "https://www.gestorias.es/tarragona/tortosa/mulet-lluis-13765"),
    ("Salvadó Assessors", "Tortosa", None, "https://www.salvadoassessors.com/"),
    ("Gestoría Online", "Tortosa", None, "https://www.gestorias.es/tarragona/tortosa/gestoria-online-17459"),
    ("Pijoan Caubet Economistes", "Tortosa", "info@caubeteconomistes.com", "https://caubeteconomistes.com/"),
    ("Jordi Pla Ferre", "Tortosa", None, "https://www.gestorias.es/tarragona/tortosa"),
    ("Gassó & Gassó Assessors", "Tortosa", "maite@gassoassessors.com", "https://www.gassoassessors.com/es/contactar"),
    ("Lexebre Consultors", "Tortosa", "info@lexebre.com", "https://www.lexebre.com/contacta/"),
    ("Tax Tortosa", "Tortosa", None, "https://www.gestorias.es/tarragona/tortosa"),
    ("Axyt Consultoria", "Tortosa", "info@axytconsultoria.com", "https://www.cylex.es/tortosa/axyt-consultoria-13499736.html"),
    ("ALC Assessors", "Tortosa", None, "https://www.alc-assessors.com/"),
    ("Mayo Consultors", "Tortosa", "despatx@mayoconsultors.com", "https://mayoconsultors.com/contacto/"),
    ("Vidiella Consultoria (Vidiella & Rosa)", "Tortosa", "auditoria@vidiellarosa.com", "https://www.vidiellarosa.com/es/contacto"),
    ("Gestoría Margalef", "Tortosa", "margalefgestoria@gmail.com", "https://www.gestoriamargalef.es/"),
    ("Rodolfo Lamote de Grignon Isuar", "Tortosa", None, "https://www.gestorias.es/tarragona/tortosa/rodolfo-lamote-de-grignon-isuar-4013"),
    ("Agest Gestió", "Tortosa", None, "https://www.gestorias.es/tarragona/tortosa"),
    ("Miravalls Assessors", "Tortosa", "miravalls@miravalls.net", "https://www.miravalls.net/"),
    ("Viñes Assessors", "Tortosa", "hola@vinyes.info", "https://vinyes.info/es/contacto/"),
    ("Assessoria Fiscal Dertusa SL", "Tortosa", None, "https://www.paginasamarillas.es/f/tortosa/assessoria-fiscal-dertusa-sl_230503856_000000001.html"),
    ("Albacar Assessors i Consultors SL", "Tortosa", "info@albacarassessors.com", "https://www.albacarassessors.com/en/"),
    ("Sánchez Medina Assessors", "Tortosa", None, "https://www.serveisactius.cat/es/baixebre/asesoria-tortosa"),
    ("Sebastià Assessors", "Tortosa", "tortosa@sebastia.info", "https://empresite.eleconomista.es/SEBASTIA-ASSESSORS.html"),
]


def emails_existentes():
    vistos = set()
    for path in ("cola_envios.xlsx", "contactos_catalan.xlsx", "contactos_castellano.xlsx",
                 "contactos_catalan_ronda2.xlsx", "contactos_ronda3.xlsx", "contactos_ronda4.xlsx"):
        try:
            wb = load_workbook(path)
        except FileNotFoundError:
            continue
        ws = wb.active
        h = [c.value for c in ws[1]]
        if "email" not in h:
            continue
        idx = h.index("email")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx]:
                vistos.add(str(row[idx]).strip().lower())
    return vistos


vistos = emails_existentes()
wb = load_workbook("cola_envios.xlsx")
ws = wb.active

nuevas = duplicadas = sin_email = 0
for empresa, ciudad, email, fuente in RONDA5:
    if not email:
        ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", "", "sin email", "", "", fuente])
        sin_email += 1
        continue
    if email.strip().lower() in vistos:
        duplicadas += 1
        continue
    vistos.add(email.strip().lower())
    ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", email, "", "", FECHA_HOY, fuente])
    nuevas += 1

wb.save("cola_envios.xlsx")
print(f"Ronda 5: {nuevas} nuevas con email (fecha {FECHA_HOY}), {duplicadas} duplicadas omitidas, {sin_email} sin email.")
