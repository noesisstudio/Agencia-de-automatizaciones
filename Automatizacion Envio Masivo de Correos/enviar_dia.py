"""Envía el lote de hoy de cola_envios.xlsx: todas las filas con fecha_prevista <= hoy
que no estén ya enviadas, hasta un máximo de --cap (por defecto 95, el límite real de
Hostinger de 100/día menos margen para poder responder tú a los clientes). Si un día se queda
corto por el límite de envíos del propio Hostinger, lo que falte se reintenta el
día siguiente sin perder nada (las filas no enviadas simplemente siguen pendientes).

Uso: .venv\\Scripts\\python.exe enviar_dia.py
"""
import argparse
import datetime
import time

import enviar_correos as ec
from openpyxl import load_workbook

ASUNTOS = {
    "catalan": "{{empresa}}, quantes hores perdeu introduint dades a mà?",
    "castellano": "{{empresa}}, ¿cuántas horas perdéis introduciendo datos a mano?",
}
PLANTILLAS = {
    "catalan": "plantilla_gestorias.html",
    "castellano": "plantilla_gestorias_es.html",
}


def main():
    parser = argparse.ArgumentParser(description="Envío diario controlado desde cola_envios.xlsx")
    parser.add_argument("--excel", default="cola_envios.xlsx")
    parser.add_argument("--cap", type=int, default=95, help="Máximo de correos a enviar hoy")
    parser.add_argument("--fecha", default=None, help="Fecha a considerar 'hoy' (AAAA-MM-DD), por defecto la fecha real")
    args = parser.parse_args()

    hoy = datetime.date.fromisoformat(args.fecha) if args.fecha else datetime.date.today()
    config = ec.cargar_config()
    plantillas_html = {idioma: open(path, encoding="utf-8").read() for idioma, path in PLANTILLAS.items()}

    wb = load_workbook(args.excel)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(headers)}

    # Direcciones ya enviadas en cualquier fila: evita mandar dos veces el mismo
    # correo si un contacto aparece duplicado en la cola.
    ya_enviados = {
        (ws.cell(row=r, column=col["email"]).value or "").strip().lower()
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=col["estado"]).value == "enviado"
    }

    server = ec.connect_smtp(config["host"], config["port"], config["user"], config["password"])
    enviados = errores = 0
    try:
        for row in range(2, ws.max_row + 1):
            if enviados >= args.cap:
                print(f"Alcanzado el tope de {args.cap} para hoy. El resto queda para el día siguiente.")
                break

            if ws.cell(row=row, column=col["estado"]).value == "enviado":
                continue

            fecha_prevista = ws.cell(row=row, column=col["fecha_prevista"]).value
            if fecha_prevista and datetime.date.fromisoformat(str(fecha_prevista)) > hoy:
                continue

            idioma = ws.cell(row=row, column=col["idioma"]).value or "catalan"
            data = {headers[i]: ws.cell(row=row, column=i + 1).value for i in range(len(headers))}
            to_addr = (data.get("email") or "").strip()
            if not to_addr:
                continue

            if to_addr.lower() in ya_enviados:
                ws.cell(row=row, column=col["estado"], value="duplicado")
                print(f"Saltado duplicado: {to_addr}")
                continue

            subject = ec.render(ASUNTOS[idioma], data)
            body = ec.render(plantillas_html[idioma], data)

            try:
                ec.enviar_email(server, config["from_email"], config["from_name"], to_addr, subject, body)
                ws.cell(row=row, column=col["estado"], value="enviado")
                ws.cell(row=row, column=col["fecha_envio"], value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
                ya_enviados.add(to_addr.lower())
                enviados += 1
                print(f"Enviado a {to_addr}")
            except Exception as e:
                if isinstance(e, __import__("smtplib").SMTPServerDisconnected):
                    wb.save(args.excel)
                    print(f"\nEl servidor cortó la conexión tras {enviados} correos. Se reintenta mañana.")
                    break
                ws.cell(row=row, column=col["estado"], value=f"error: {e}")
                errores += 1
                print(f"Error enviando a {to_addr}: {e}")

            wb.save(args.excel)
            time.sleep(config["delay"])
    finally:
        try:
            server.quit()
        except Exception:
            pass
        wb.save(args.excel)

    print(f"\nResumen del día {hoy.isoformat()}: {enviados} enviados, {errores} errores.")


if __name__ == "__main__":
    main()
