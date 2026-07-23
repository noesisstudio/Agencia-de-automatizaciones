"""Añade a cola_envios.xlsx la ronda 6 (Badalona, L'Hospitalet de Llobregat,
Santa Coloma de Gramenet — búsqueda 17/07/2026), deduplicando contra todos los
emails ya usados. Después ejecutar: reconstruir_maestro.py"""
from openpyxl import load_workbook

FECHA_HOY = "2026-07-17"

RONDA6 = [
    ("Assessoria Arribas", "Badalona", "info@asesoriaarribas.com", "https://asesoriaarribas.com/"),
    ("Gestoria CR (CR Grup)", "Badalona", "info@crgrup.com", "http://www.crgrup.com/"),
    ("Tax 360", "Badalona", "hola@tax360.es", "http://www.tax360.es/"),
    ("Fin2Go", "Badalona", "info@fin2go.es", "https://fin2go.es/"),
    ("Interquatre Consultors", "Badalona", "info@interquatre.com", "https://www.interquatre.com/"),
    ("Gestoría Montserrat", "Badalona", "gestoria@gestoriamontserrat.com", "https://gestoriamontserrat.com/"),
    ("Open Tax and Legal", "Badalona", "info@opentaxandlegal.com", "https://www.opentaxandlegal.com/"),
    ("Sires Consulting", "Badalona", "info@siresconsulting.es", "https://www.siresconsulting.es/"),
    ("Gestoría Ruppmann", "Badalona", "info@gestoriaruppmann.es", "https://gestoriaruppmann.com/"),
    ("Assessoria Asiz", "Badalona", "info@asiz.es", "https://asiz.es/"),
    ("SGI 360 SL", "Badalona", "info@sgi360.es", "https://www.sgi360.es/"),
    ("ps tràmits", "Badalona", "pstramits@pstramits.com", "https://www.pstramits.es/"),
    ("Asesoría Jurídico Laboral Suárez", "Badalona", "contacto@ajlsuarez.com", "https://ajlsuarez.com/"),
    ("García Castro Asesores SLP", "Badalona", "info@garciacastroasesores.com", "https://www.garciacastroasesores.com/"),
    ("Asesoria Gayse SL", "Badalona", "agayse@asesoriagayse.es", "https://asesoriagayse.es/"),
    ("Gabinet Torres Assessors SL", "Badalona", "mail@gestoriatorres.com", "https://www.gestoriatorres.com/"),
    ("Campos Asesores", "Badalona", "info@camposasesores.es", "https://www.camposasesores.es/"),
    ("Asesoría Vázquez", "Badalona", "info@asesoria-vazquez.es", "https://asesoria-vazquez.es/"),
    ("Alfet SL", "Badalona", "alfet@alfet.es", "https://alfet.es/"),
    ("Pérez Madrid Asesores SL", "Badalona", "administracion@perezmadrid.com", "https://www.perezmadrid.com/contacto/"),
    ("Asesoría Arrahona (EREL Gestoría)", "Badalona", "erel@erelgestiononline.com", "https://www.erelgestiononline.com/"),
    ("Lladó Grup Consultor", "Badalona", None, "https://www.lladogrup.com/contacte/"),
    ("Sayol Assessors", "Badalona", None, "https://www.gestorias.es/barcelona/badalona"),
    ("Asesoría Pereira", "Badalona", None, "https://www.gestorias24.com/ubicaciones/badalona"),

    ("Grup Pareto", "L'Hospitalet de Llobregat", "pareto@gruppareto.com", "https://gruppareto.com/"),
    ("TF Assessors", "L'Hospitalet de Llobregat", "info@tfassessors.com", "https://tfassessors.com/"),
    ("PyG Asesoría", "L'Hospitalet de Llobregat", "info@pygasesoria.com", "https://pygasesoria.com/"),
    ("Gestoría Ruiz", "L'Hospitalet de Llobregat", "gestoria@gestoria-ruiz.es", "http://www.gestoria-ruiz.es/"),
    ("Vilay Consultores", "L'Hospitalet de Llobregat", "informa@vilayconsultores.com", "https://vilayconsultores.com/"),
    ("Gonzalez & Asociados", "L'Hospitalet de Llobregat", "info@gonzalezasociadosbcn.es", "https://www.gonzalezasociadosbcn.es/"),
    ("Gespolcat", "L'Hospitalet de Llobregat", "info@gestorhospitalet.es", "https://www.gestorhospitalet.es/contacto/"),
    ("Issei Asesoria", "L'Hospitalet de Llobregat", "issei@isseiasesoria.com", "https://www.asesorias-empresas.es/directorio-asesorias/barcelona/lhospitalet-de-llobregat/issei-asesoria/"),
    ("Moral i Associats", "L'Hospitalet de Llobregat", "info@moraliassociats.com", "http://www.moraliassociats.com/contacto/"),
    ("Gestoría Hospitalet", "L'Hospitalet de Llobregat", "contacto@gestoriahospitalet.es", "https://gestoriahospitalet.es/contacto"),
    ("Bravo Reinon", "L'Hospitalet de Llobregat", "info@bravoreinon.com", "https://bravoreinon.com/contacto/"),
    ("Herasol Assessors", "L'Hospitalet de Llobregat", "herasol@herasolassessors.com", "https://herasolassessors.es/contacto/"),
    ("Asesoria la Torrassa (Gestions i Serveis)", "L'Hospitalet de Llobregat", "info@gestionsiserveis.es", "https://www.gestionsiserveis.es/"),
    ("EIO Asesores", "L'Hospitalet de Llobregat", "info@eioasesores.com", "https://eioasesores.com/contacto/"),
    ("Gestoria Plus", "L'Hospitalet de Llobregat", "info@gestoriaplus.es", "https://gestoriaplus.es/contacto/"),
    ("Assessoria Vilagestió", "L'Hospitalet de Llobregat", "info@vilagestio.com", "https://vilagestio.com/contacto-asesoria-vilagestio/"),
    ("Borrell Gestora", "L'Hospitalet de Llobregat", None, "https://www.borrellgestora.com/contacto/"),
    ("Breakeven Assessors (Culebras)", "L'Hospitalet de Llobregat", None, "https://www.culebras.es/contacto/"),
    ("HV Assessors", "L'Hospitalet de Llobregat", None, "https://hvasesores.com/"),
    ("Ormatiz Assessors SL", "L'Hospitalet de Llobregat", None, "https://empresite.eleconomista.es/ORMATIZ-ASSESSORS.html"),
    ("Gestoría Terrazas Bcn SLP", "L'Hospitalet de Llobregat", None, "https://www.gestorias24.com/ubicaciones/l-hospitalet-de-llobregat"),

    ("M&S Asesores", "Santa Coloma de Gramenet", "jisern@mys.cat", "https://www.mys.cat/"),
    ("Rubio Gestors Administratius SL", "Santa Coloma de Gramenet", "gestoriarubio@rubiogestoria.com", "https://rubiogestoria.com/"),
    ("Tacer Gestió Administrativa", "Santa Coloma de Gramenet", "tacer@tacergestion.com", "https://tacergestion.com.es/"),
    ("Gestoría Asesoría Martín", "Santa Coloma de Gramenet", "ruben.martin@asesoriamartin.com", "https://gestoria-martin.com/"),
    ("GrupLegal", "Santa Coloma de Gramenet", "gruplegal@gruplegal.com", "https://gruplegal.com/contacto/"),
    ("Assessories Nàpols SL", "Santa Coloma de Gramenet", "info@asesoriasnapoles.es", "https://asesoriasnapoles.es/"),
    ("Gestoria SALEC", "Santa Coloma de Gramenet", "msalec@gestoriasalec.com", "https://gestoriasalec.com/"),
    ("Gestoría Global Gestors SLP", "Santa Coloma de Gramenet", "info@globalgestors.com", "https://globalgestors.com/"),
    ("Demos Advocats & Assessors", "Santa Coloma de Gramenet", "info@demosadvocats.com", "https://demosadvocats.com/"),
    ("Asesoría Pellicer", "Santa Coloma de Gramenet", "correo@asesoriapellicer.es", "https://asesoriapellicer.es/"),
    ("Gestoría Eurogestió", "Santa Coloma de Gramenet", "eurogestion@eurogestionsl.com", "https://eurogestionsl.com/"),
    ("Esfera Assessors", "Santa Coloma de Gramenet", "laboral@esferaassessors.es", "https://esferaassessors.es/"),
    ("Assegurances Carulla - Gestoria Cano", "Santa Coloma de Gramenet", "gestoria@gtcac.net", "https://gtcac.cat/"),
    ("Gestoría Arce y López SL", "Santa Coloma de Gramenet", "gestoria@arcelopez.com", "http://www.gestoriaarceylopez.com/"),
    ("Gestoría Rivero", "Santa Coloma de Gramenet", "info@riverogestoria.com", "https://www.riverogestoria.com/contacto/contacto/ia37"),
    ("Grupo Tramitalia", "Santa Coloma de Gramenet", "info@grupotramitalia.com", "https://empresite.eleconomista.es/TRAMITALIA-TDA-CONSULTORIA-GESTION-INTEGRAL.html"),
    ("Aseminse SL", "Santa Coloma de Gramenet", None, "https://www.paginasamarillas.es/f/santa-coloma-de-gramenet/aseminse-s-l-_202599965_000000002.html"),
    ("Juridical Business SLP", "Santa Coloma de Gramenet", None, "https://juridicalbusiness.es/"),
    ("Gestoria Dat@", "Santa Coloma de Gramenet", None, "https://www.gestorias24.com/ubicaciones/santa-coloma-de-gramenet"),
    ("Asesoría Asgar", "Santa Coloma de Gramenet", None, "http://www.asgar.es/"),
    ("Andreu Assessors", "Santa Coloma de Gramenet", None, "https://www.paginasamarillas.es/f/santa-coloma-de-gramenet/andreu-assessors_009459256_000000001.html"),
]


def emails_existentes():
    vistos = set()
    for path in ("cola_envios.xlsx", "contactos_catalan.xlsx", "contactos_castellano.xlsx",
                 "contactos_catalan_ronda2.xlsx", "contactos_ronda3.xlsx", "contactos_ronda4.xlsx"):
        try:
            wb = load_workbook(path)
        except FileNotFoundError:
            continue
        ws = wb.active
        h = [c.value for c in ws[1]]
        if "email" not in h:
            continue
        idx = h.index("email")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx]:
                vistos.add(str(row[idx]).strip().lower())
    return vistos


vistos = emails_existentes()
wb = load_workbook("cola_envios.xlsx")
ws = wb.active

nuevas = duplicadas = sin_email = 0
for empresa, ciudad, email, fuente in RONDA6:
    if not email:
        ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", "", "sin email", "", "", fuente])
        sin_email += 1
        continue
    if email.strip().lower() in vistos:
        duplicadas += 1
        continue
    vistos.add(email.strip().lower())
    ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", email, "", "", FECHA_HOY, fuente])
    nuevas += 1

wb.save("cola_envios.xlsx")
print(f"Ronda 6: {nuevas} nuevas con email (fecha {FECHA_HOY}), {duplicadas} duplicadas omitidas, {sin_email} sin email.")
