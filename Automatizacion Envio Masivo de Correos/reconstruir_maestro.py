"""Reconstruye maestro_gestorias.xlsx desde los archivos fuente + cola_envios.xlsx,
PRESERVANDO los comentarios manuales y las marcas de 'rebotado' del maestro actual.
Ejecutar siempre este script (y no versiones anteriores) para regenerar el maestro."""
from collections import Counter

from openpyxl import Workbook, load_workbook

# --- preservar comentarios y rebotes del maestro actual ---
comentarios = {}
rebotados = set()
try:
    wbm = load_workbook("maestro_gestorias.xlsx")
    wsm = wbm.active
    hm = [c.value for c in wsm[1]]
    cm = {n: i for i, n in enumerate(hm)}
    for row in wsm.iter_rows(min_row=2, values_only=True):
        if "comentario" in cm and row[cm["comentario"]]:
            comentarios[(row[cm["empresa"]], row[cm["ciudad"]])] = row[cm["comentario"]]
        if row[cm["estado"]] == "rebotado" and row[cm["email"]]:
            rebotados.add(str(row[cm["email"]]).strip().lower())
except FileNotFoundError:
    pass

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "maestro"
ws_out.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio",
               "fecha_prevista", "fuente_email", "comentario"])


def fila(nombre, empresa, ciudad, idioma, email, estado, fecha_envio, fecha_prev, fuente):
    if email and str(email).strip().lower() in rebotados:
        estado = "rebotado"
    ws_out.append([nombre, empresa, ciudad, idioma, email or "", estado,
                   fecha_envio or "", fecha_prev or "", fuente or "",
                   comentarios.get((empresa, ciudad), "")])


def copiar_desde(path, excluir_pendientes_con_email=False):
    wbx = load_workbook(path)
    wsx = wbx.active
    hx = [c.value for c in wsx[1]]
    cx = {n: i for i, n in enumerate(hx)}
    for row in wsx.iter_rows(min_row=2, values_only=True):
        estado = row[cx["estado"]]
        email = row[cx["email"]]
        if excluir_pendientes_con_email and estado != "enviado" and email:
            continue
        ef = "enviado" if estado == "enviado" else ("sin email" if not email else "pendiente")
        fila(row[cx["nombre"]], row[cx["empresa"]], row[cx["ciudad"]],
             "castellano" if "castellano" in path else "catalan",
             email, ef, row[cx["fecha_envio"]] if estado == "enviado" else "", "",
             row[cx["fuente_email"]] if "fuente_email" in cx else "")


copiar_desde("contactos_catalan.xlsx")
copiar_desde("contactos_castellano.xlsx")
copiar_desde("contactos_catalan_ronda2.xlsx", excluir_pendientes_con_email=True)

wbc = load_workbook("cola_envios.xlsx")
wsc = wbc.active
hc = [c.value for c in wsc[1]]
cc = {n: i for i, n in enumerate(hc)}
for row in wsc.iter_rows(min_row=2, values_only=True):
    if row[cc["estado"]] == "enviado":
        ef = "enviado"
    elif row[cc["estado"]] == "duplicado":
        ef = "duplicado"
    elif not row[cc["email"]]:
        ef = "sin email"
    else:
        ef = "pendiente"
    fila(row[cc["nombre"]], row[cc["empresa"]], row[cc["ciudad"]], row[cc["idioma"]],
         row[cc["email"]], ef, row[cc["fecha_envio"]], row[cc["fecha_prevista"]], row[cc["fuente_email"]])

wb_out.save("maestro_gestorias.xlsx")
cnt = Counter()
for row in ws_out.iter_rows(min_row=2, values_only=True):
    cnt[row[5]] += 1
print("maestro reconstruido:", dict(cnt))
print("comentarios preservados:", len(comentarios))
print("rebotados preservados:", len(rebotados))
