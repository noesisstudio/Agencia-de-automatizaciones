"""Añade a cola_envios.xlsx la ronda 7 (Gavà, Castelldefels, Esplugues de Llobregat
— búsqueda 20/07/2026), deduplicando contra todos los emails ya usados. Después
ejecutar: reconstruir_maestro.py"""
from datetime import date, timedelta

from openpyxl import load_workbook

FECHA_HOY = date.today().isoformat()

RONDA7 = [
    ("Miret Consultoria Jurídica", "Gavà", None, "https://www.gestorias.es/barcelona/gava"),
    ("Sist Consulting (Silvia Salmerón)", "Gavà", None, "https://www.gestorias.es/barcelona/gava"),
    ("Esica Consulting", "Gavà", None, "https://www.gestorias.es/barcelona/gava/esica-consulting-19341"),
    ("Mcr Consultores", "Gavà", "mcr@mcrgestoriaonline.com", "https://www.mcrgestoriaonline.com/contacto/"),
    ("Tax Gavà", "Gavà", "central@tax.es", "https://www.tax.es/es/oficinas/cataluna/tax-gava.html"),
    ("Murillo Gestió Integral", "Gavà", "murillogestio@murillogestio.es", "https://empresite.eleconomista.es/MURILLO-GESTIO-INTEGRAL.html"),
    ("Cristina Gil Higuera Gestores", "Gavà", None, "https://www.gestorias.es/barcelona/gava/cristina-gil-higuera-gestores-905"),
    ("J.R. Marín Assessors", "Gavà", "jrmarin@icab.es", "https://www.qdq.com/jr-marin-assessors-314246"),
    ("Gestilabor Cithe", "Gavà", None, "https://www.gestorias.es/barcelona/gava/gestilabor-cithe-4860"),
    ("Alemany i Tobia Associats", "Gavà", None, "https://www.gestorias.es/barcelona/gava"),
    ("Geudigest Integral", "Gavà", None, "https://www.gestorias.es/barcelona/gava/geudigest-integral-9616"),
    ("Gestoria Tarrida", "Gavà", "assessoria@tarrida.com", "https://www.tarrida.com/es/contacto/"),
    ("Asesoría Navarro", "Gavà", "info@asesorianavarro.es", "https://asesorianavarro.es/contacto/"),
    ("Consultoria i Assessoria Bonet S.L.P.", "Gavà", None, "https://empresite.eleconomista.es/CONSULTORIA-ASSESSORIA-BONET-SLP.html"),
    ("Asesoria Mundial de Transportes S.L.", "Gavà", None, "https://empresite.eleconomista.es/Actividad/ASESORIA-GESTORIA/localidad/GAVA-BARCELONA/"),
    ("Asesoria de Servicios Sitges S.L.", "Gavà", None, "https://empresite.eleconomista.es/Actividad/ASESORIA-GESTORIA/localidad/GAVA-BARCELONA/"),
    ("Lex Asociados - Asesoria Fislab S.L.", "Gavà", "info@asesoriafislab.com", "https://www.asesoriafislab.com/"),
    ("Isigma Asesoria Tecnológica S.L.", "Gavà", None, "https://empresite.eleconomista.es/Actividad/ASESORIA-GESTORIA/localidad/GAVA-BARCELONA/"),
    ("Management Consultancy Review S.L.", "Gavà", None, "https://empresite.eleconomista.es/Actividad/ASESORIA-GESTORIA/localidad/GAVA-BARCELONA/"),
    ("Gestibaix", "Gavà", None, "https://www.paginasamarillas.es/f/gava/gestibaix_200647832_000000001.html"),
    ("Gorriz - Arias Consulting (Gavà)", "Gavà", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/gava/"),
    ("Emilio García Ramos", "Gavà", "consultorip@emiliogarciaweb.com", "http://www.emiliogarciaweb.es/"),
    ("Izquierdo i Tugas Associats", "Gavà", "info@gestoria-izquierdo.cat", "https://gestoria-izquierdo.cat/contactar-amb-izquierdo-i-tugas-associats/"),
    ("Grupo Asesor y de Servicios Generales", "Gavà", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/gava/"),
    ("Asem Visiones Competitivas S.L.", "Gavà", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/gava/"),
    ("Asom Asesoría", "Gavà", None, "https://asomasesoria.com/"),
    ("Ibáñez Estudio Asesoría S.L.", "Gavà", None, "https://empresite.eleconomista.es/Actividad/ASESORIA-GESTORIA/localidad/GAVA-BARCELONA/"),

    ("ALR - Asesoría laboral, fiscal y contable", "Castelldefels", "info@alrasesoria.com", "https://alrasesoria.com/contacto/"),
    ("ANTEO Asesoría Castelldefels", "Castelldefels", None, "https://anteo.es/"),
    ("Jurojin Consulting / Huerga y Asociados", "Castelldefels", "huergayasociados@huergayasociados.com", "https://huergayasociados.com/contacto/"),
    ("Asesoría Mafer", "Castelldefels", "administracion@asesoriamafer.com", "https://asesoriamafer.com/"),
    ("BYR Assessors", "Castelldefels", "byr@byr.cat", "https://www.byr.cat/"),
    ("Ortega&Cao Assessors", "Castelldefels", None, "https://ortegacaoassessors.com/contacto/"),
    ("EH Asesores", "Castelldefels", None, "https://www.ehasesores.com/contacto/"),
    ("Tarves Consultores", "Castelldefels", "info@asesoriacastelldefels.es", "https://asesoriacastelldefels.es/"),
    ("Parentesi Consultoria Global S.L.", "Castelldefels", "info@parentesiconsultoria.com", "https://parentesiconsultoria.com/contacto/"),
    ("Servigest", "Castelldefels", "servigest@servigestsccl.es", "https://servigestsccl.es/"),
    ("Subias Economistas", "Castelldefels", None, "https://www.subiaseconomistas.com/"),
    ("IPE Asesoría", "Castelldefels", "ipeasesoria@ipeasesoria.com", "https://www.ipeasesoria.com/"),
    ("Assessoria Roca", "Castelldefels", "info@assessoriaroca.com", "https://www.assessoriaroca.com/contacto/"),
    ("Gestión Eficaz", "Castelldefels", "juridico@gestioneficaz.net", "https://gestioneficaz.net/contacto/"),
    ("G.M. Fiscomer", "Castelldefels", "manuel@fiscomer.net", "https://empresite.eleconomista.es/GM-FISCOMER.html"),
    ("Centro Gestor Javier", "Castelldefels", "centrogestor10@gmail.com", "https://www.gestorias.es/barcelona/castelldefels/centro-gestor-javier-817"),
    ("Assessors 2000", "Castelldefels", "a2000@assessors2000.com", "https://assessors2000.com/asesoria-gestoria/"),
    ("Fels Consulting 2005", "Castelldefels", "mbustos@felsconsulting.com", "https://gestoriacastelldefels.es/"),
    ("Asifb - Gestoría Castelldefels", "Castelldefels", None, "https://www.asifb.com/"),
    ("Gascón Abogados", "Castelldefels", None, "http://www.bufetegascon.com/contacto-asesoria-gestoria-castelldefels/"),

    ("T&B Assessors", "Esplugues de Llobregat", "despacho@tbassessors.es", "https://www.tbassessors.es/"),
    ("Gavaldà Associats", "Esplugues de Llobregat", None, "https://www.gestorias.es/barcelona/esplugues-de-llobregat/gavalda-associats-16543"),
    ("Gestoría Molina & Giró", "Esplugues de Llobregat", None, "https://www.gestorias.es/barcelona/esplugues-de-llobregat/silvia-molina-salmeron-16356"),
    ("Asesoría Esplugues", "Esplugues de Llobregat", "asesoria@asesoriaesplugues.com", "https://www.asesoriaesplugues.com/contacto/"),
    ("Nil i Martí", "Esplugues de Llobregat", None, "https://www.gestorias.es/barcelona/esplugues-de-llobregat/nil-i-marti-18840"),
    ("Issa Assessorament i Gestió", "Esplugues de Llobregat", None, "http://www.issasl.com/"),
    ("Asesoría JL", "Esplugues de Llobregat", None, "https://www.paginasamarillas.es/f/esplugues-de-llobregat/asesoria-j-l-_196921530_000000001.html"),
    ("Castellà Fontes Assessors", "Esplugues de Llobregat", "info@castella-bcn.com", "https://www.castella-bcn.com/en/contact/"),
    ("Fátima Zohra Bradhi", "Esplugues de Llobregat", None, "https://www.gestorias.es/barcelona/esplugues-de-llobregat/fatima-zohra-bradhi-2855"),
    ("Gestoría Giménez Casero & Asociados", "Esplugues de Llobregat", None, "https://www.gestorias.es/barcelona/esplugues-de-llobregat/gestoria-gimenez-casero-asociados-4518"),
    ("Felip García Associats", "Esplugues de Llobregat", None, "https://www.gestorias.es/barcelona/esplugues-de-llobregat/felip-garcia-associats-6485"),
    ("Culebras Assessors", "Esplugues de Llobregat", "culebras@culebras.es", "https://empresite.eleconomista.es/CULEBRAS-ASSESSORS.html"),
    ("E.R. Assessors", "Esplugues de Llobregat", "info@erassessors.com", "https://www.erassessors.com/es/"),
    ("Gremicat (oficina Esplugues)", "Esplugues de Llobregat", "gremicat@gremicat.es", "https://www.gremicat.es/gestoria-esplugues-de-llobregat/"),
    ("JOB Asesoria y Gestión S.L.", "Esplugues de Llobregat", "dordonez@jobauditores.com", "https://www.northdata.com/Job%20Asesoria%20y%20Gestion%20SL,%20Esplugues%20de%20Llobregat/NIF%20B60502507"),
    ("Serimar Esplugues S.L.", "Esplugues de Llobregat", "f802srm@outlook.es", "https://empresite.eleconomista.es/SERIMAR-ESPLUGAS.html"),
    ("Esplugest S.L.", "Esplugues de Llobregat", "mdonaire@esplugest.es", "https://empresite.eleconomista.es/ESPLUGEST.html"),
    ("Asesoría J. Samitier", "Esplugues de Llobregat", "samitier@jsamitier.com", "https://jsamitier.com/contacto-asesoria-jsamitier-esplugues-llobregat/"),
    ("CE Consulting Esplugues", "Esplugues de Llobregat", None, "https://ceconsulting.es/oficina/barcelona-esplugues-de-llobregat/"),
    ("Adela Asesores S.L. (Esplugues)", "Esplugues de Llobregat", "info@adelaasesores.com", "https://www.adelaasesores.com/ca/contacte/"),
    ("Coper Press S.L.", "Esplugues de Llobregat", None, "https://www.paginasamarillas.es/a/asesorias-laborales/barcelona/esplugues-de-llobregat/"),
]


def emails_existentes():
    vistos = set()
    for path in ("cola_envios.xlsx", "contactos_catalan.xlsx", "contactos_castellano.xlsx",
                 "contactos_catalan_ronda2.xlsx", "contactos_ronda3.xlsx", "contactos_ronda4.xlsx"):
        try:
            wb = load_workbook(path)
        except FileNotFoundError:
            continue
        ws = wb.active
        h = [c.value for c in ws[1]]
        if "email" not in h:
            continue
        idx = h.index("email")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx]:
                vistos.add(str(row[idx]).strip().lower())
    return vistos


vistos = emails_existentes()
wb = load_workbook("cola_envios.xlsx")
ws = wb.active

nuevas = duplicadas = sin_email = 0
for empresa, ciudad, email, fuente in RONDA7:
    if not email:
        ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", "", "sin email", "", "", fuente])
        sin_email += 1
        continue
    if email.strip().lower() in vistos:
        duplicadas += 1
        continue
    vistos.add(email.strip().lower())
    ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", email, "", "", FECHA_HOY, fuente])
    nuevas += 1

wb.save("cola_envios.xlsx")
print(f"Ronda 7: {nuevas} nuevas con email (fecha {FECHA_HOY}), {duplicadas} duplicadas omitidas, {sin_email} sin email.")
