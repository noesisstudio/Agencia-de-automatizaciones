"""Sincroniza los emails que el usuario ha añadido a mano en maestro_gestorias.xlsx
(sobre filas que estaban 'sin email') de vuelta a cola_envios.xlsx, reparte de nuevo
fecha_prevista (95/día, preservando el orden ya existente para no mover lo ya
programado para hoy), y reconstruye maestro_gestorias.xlsx."""
import datetime
from openpyxl import Workbook, load_workbook

DAILY_CAP = 95
PRIMER_DIA = datetime.date(2026, 7, 14)

wb_m = load_workbook("maestro_gestorias.xlsx")
ws_m = wb_m.active
hm = [c.value for c in ws_m[1]]
cm = {n: i for i, n in enumerate(hm)}

nuevos_emails = {}
for row in ws_m.iter_rows(min_row=2, values_only=True):
    if row[cm["estado"]] == "sin email" and row[cm["email"]] and "@" in str(row[cm["email"]]):
        nuevos_emails[row[cm["empresa"]]] = row[cm["email"]].strip()

wb = load_workbook("cola_envios.xlsx")
ws = wb.active
h = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(h)}

actualizadas = []
for r in range(2, ws.max_row + 1):
    empresa = ws.cell(row=r, column=col["empresa"]).value
    if empresa in nuevos_emails and not ws.cell(row=r, column=col["email"]).value:
        ws.cell(row=r, column=col["email"], value=nuevos_emails[empresa])
        ws.cell(row=r, column=col["estado"], value="")
        actualizadas.append(empresa)

pendientes_rows = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(row=r, column=col["estado"]).value != "enviado"
                   and ws.cell(row=r, column=col["email"]).value]

for i, r in enumerate(pendientes_rows):
    dia = PRIMER_DIA + datetime.timedelta(days=i // DAILY_CAP)
    ws.cell(row=r, column=col["fecha_prevista"], value=dia.isoformat())

wb.save("cola_envios.xlsx")
print(f"Sincronizadas {len(actualizadas)} gestorías con email nuevo: {actualizadas}")

# --- reconstruir maestro_gestorias.xlsx ---
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "maestro"
ws_out.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio",
               "fecha_prevista", "fuente_email", "comentario"])


def copiar_desde(path, excluir_pendientes_con_email=False):
    wbx = load_workbook(path)
    wsx = wbx.active
    hx = [c.value for c in wsx[1]]
    cx = {n: i for i, n in enumerate(hx)}
    for row in wsx.iter_rows(min_row=2, values_only=True):
        estado = row[cx["estado"]]
        email = row[cx["email"]]
        if excluir_pendientes_con_email and estado != "enviado" and email:
            continue
        estado_final = "enviado" if estado == "enviado" else ("sin email" if not email else "pendiente")
        ws_out.append([
            row[cx["nombre"]], row[cx["empresa"]], row[cx["ciudad"]],
            "castellano" if "castellano" in path else "catalan",
            email or "", estado_final, row[cx["fecha_envio"]] if estado == "enviado" else "",
            "", row[cx["fuente_email"]] if "fuente_email" in cx else "", "",
        ])


copiar_desde("contactos_catalan.xlsx")
copiar_desde("contactos_castellano.xlsx")
copiar_desde("contactos_catalan_ronda2.xlsx", excluir_pendientes_con_email=True)

wb_cola = load_workbook("cola_envios.xlsx")
ws_cola = wb_cola.active
hc = [c.value for c in ws_cola[1]]
cc = {n: i for i, n in enumerate(hc)}
for row in ws_cola.iter_rows(min_row=2, values_only=True):
    if row[cc["estado"]] == "enviado":
        estado_final = "enviado"
    elif not row[cc["email"]]:
        estado_final = "sin email"
    else:
        estado_final = "pendiente"
    ws_out.append([
        row[cc["nombre"]], row[cc["empresa"]], row[cc["ciudad"]], row[cc["idioma"]],
        row[cc["email"]] or "", estado_final, row[cc["fecha_envio"]] or "", row[cc["fecha_prevista"]] or "",
        row[cc["fuente_email"]] or "", "",
    ])

wb_out.save("maestro_gestorias.xlsx")
print(f"maestro_gestorias.xlsx reconstruido con {ws_out.max_row - 1} filas.")
