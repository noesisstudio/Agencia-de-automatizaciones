"""1) Añade las 6 ciudades nuevas (Mollet, Cornellà, Sant Boi, El Prat, Viladecans,
Molins de Rei) a cola_envios.xlsx, continuando el reparto de fecha_prevista a
DAILY_CAP/día. 2) Construye maestro_gestorias.xlsx: TODAS las gestorías vistas hasta
ahora (enviadas, pendientes con fecha prevista, y sin email encontrado) en un único
documento, con una columna 'comentario' en blanco para anotaciones manuales."""
import datetime
from openpyxl import Workbook, load_workbook

DAILY_CAP = 105

# (empresa, ciudad, email_o_None, fuente)
NUEVAS = [
    ("Gestoría Viaplana (Viaplana Multigestió)", "Mollet del Vallès", None, "https://www.gestorias.es/barcelona/mollet-del-valles/gestoria-viaplana-470"),
    ("Assessoria i Serveis Empresarials Costa", "Mollet del Vallès", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Asesoría del Vallès", "Mollet del Vallès", "info@asesoriavalles.es", "http://asesoriavalles.es/"),
    ("Lex Valles Associats", "Mollet del Vallès", "info@lexvalles.com", "https://lexvalles.es/contacto-asesoria-en-mollet/"),
    ("Fortuny i Janoher", "Mollet del Vallès", "info@fortuny-janoher.com", "http://www.fortuny-janoher.com"),
    ("AISM S.L.", "Mollet del Vallès", "aism@asesoria-aism.com", "https://aismasesoria.com/"),
    ("Gabinet Emiliano i Associats, S.L.", "Mollet del Vallès", "info@gabinetemiliano.com", "https://www.gabineteemiliano.es/"),
    ("AR Asesores Mollet", "Mollet del Vallès", None, "https://www.gestoriavalles.es/contacto/"),
    ("GCA Asesoría", "Mollet del Vallès", None, "https://www.paginasamarillas.es/f/mollet-del-valles/gca_001804467_000000001.html"),
    ("Gypesa", "Mollet del Vallès", None, "https://www.gestorias.es/barcelona/mollet-del-valles/gypesa-9131"),
    ("Gestoría Vallès", "Mollet del Vallès", None, "https://www.gestorias.es/barcelona/mollet-del-valles/gestoria-valles-17920"),
    ("Valles Gestió", "Mollet del Vallès", None, "https://www.gestorias.es/barcelona/mollet-del-valles/valles-gestio-16561"),
    ("Asesoría Gala (Gala S.C.P.)", "Mollet del Vallès", "gala@asesoriagala.com", "https://asesoriagala.com/contacto/"),
    ("Dynamic Advisers", "Mollet del Vallès", "info@dynamicadvisers.com", "https://www.dynamicadvisers.com/gestoria-mollet-del-valles-dynamic-advisers/"),
    ("Poch Assessors", "Mollet del Vallès", None, "https://www.paginasamarillas.es/f/mollet-del-valles/poch-assessors_020831442_000000003.html"),
    ("Mayolas Assessors d'Empreses", "Mollet del Vallès", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Diperex S.L.", "Mollet del Vallès", None, "https://www.einforma.com/informacion-empresa/diperex-slp"),
    ("Bemtronic Online S.L.", "Mollet del Vallès", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Assessoria Joan Mercadal S.L.", "Mollet del Vallès", None, "https://www.citiservi.es/barcelona/assessoria-joan-mercadal-mollet-del-valles__930002_67.html"),
    ("Salvador Lopez Molina", "Mollet del Vallès", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Rosa Bravo López", "Mollet del Vallès", None, "https://www.paginasamarillas.es/a/asesorias-laborales/barcelona/mollet-del-valles/"),
    ("Ecofinancial Group Consulting 86 S.L.", "Mollet del Vallès", None, "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),

    ("Gestoria Romo", "Cornellà de Llobregat", "gestoria@gestoriaromo.com", "http://www.gestoriaromo.com/contactar-con-gestoria-romo-en-cornella-de-llobregat"),
    ("Seci Asesores", "Cornellà de Llobregat", "seciasesores@seci.biz", "https://www.seciasesores.es/"),
    ("Gestoría GesRosas", "Cornellà de Llobregat", "info@gesrosas.es", "https://www.cylex.es/cornella-de-llobregat/gestor%C3%ADa-gesrosas-13269099.html"),
    ("Baix Asesoramiento y Gestión S.L.", "Cornellà de Llobregat", "consultas@asesoriabaix.com", "https://www.asesoriabaix.es/contacto/"),
    ("Gestoría Palacios", "Cornellà de Llobregat", "info@gpalacios.com", "https://gestoriapalacios.es/"),
    ("SEVIC Asesoría", "Cornellà de Llobregat", "info@asesoriasevic.com", "https://asesoriasevic.com/"),
    ("Esconfi", "Cornellà de Llobregat", "esconfi@esconfi.es", "https://esconfi.es/"),
    ("Assebaix", "Cornellà de Llobregat", "anna@assebaix.com", "https://www.assebaix.com/"),
    ("La Asesoría (Cornellà)", "Cornellà de Llobregat", "laasesoria.documents9@gmail.com", "https://laasesoriacornella.com/la-asesoria/"),
    ("Asesoría Taxperts", "Cornellà de Llobregat", "comercial@taxperts.es", "https://www.taxperts.es/contacto/"),
    ("Fiscount Tax & Accounting S.L.", "Cornellà de Llobregat", "info@fiscount.es", "https://fiscount.weebly.com/contacto.html"),
    ("Assellob", "Cornellà de Llobregat", "info@assellob.net", "https://assellob.net/contacto/"),
    ("TG Assessoria Fiscal", "Cornellà de Llobregat", "tgassessoria@tgassessoria.com", "https://www.qdq.com/t-g-assesoria-fiscal-scp-826098"),
    ("Anteo Asesoría en Cornellà", "Cornellà de Llobregat", "info@anteoetl.com", "https://anteo.es/en/contact/"),
    ("Agestem S.L.", "Cornellà de Llobregat", "asesoria@agestem.com", "https://www.agestem.com/"),
    ("J.M. Medina S.L.", "Cornellà de Llobregat", None, "https://www.gestorias.es/barcelona/cornella-de-llobregat/j-m-medina-1673"),
    ("Moreno Salcedo Abogados y Economistas", "Cornellà de Llobregat", "info@morenosalcedo.com", "https://morenosalcedo.com/"),
    ("Daspime S.L.", "Cornellà de Llobregat", "bego@daspime.net", "https://www.paginasamarillas.es/f/cornella-de-llobregat/daspime-s-l_233001569_000000001.html"),
    ("Asesoría Roser Camps S.L.", "Cornellà de Llobregat", None, "https://www.paginasamarillas.es/f/cornella-de-llobregat/asesoria-roser-camps-s-l-_014410344_000000001.html"),
    ("Gestión y Asesoramientos Cornellà", "Cornellà de Llobregat", None, "https://www.gestorias.es/barcelona/cornella-de-llobregat/gestion-y-asesoramientos-cornella-9711"),
    ("D.M. Asesoría", "Cornellà de Llobregat", None, "https://firmania.es/cornella-de-llobregat/dm-asesor%C3%ADa-1693298"),
    ("Berule Consult S.L.", "Cornellà de Llobregat", None, "https://firmania.es/cornella-de-llobregat/berule-consult-sl-1797127"),

    ("Assessoria i Gestió Negre", "Sant Boi de Llobregat", "info@assessorianegre.com", "https://assessorianegre.com/contacto/"),
    ("ARSA Gestión", "Sant Boi de Llobregat", "info@arsagestion.com", "https://www.arsagestion.com/asesoria-sant-boi"),
    ("Assessoria Mendiola", "Sant Boi de Llobregat", "amendiola@amendiola.es", "https://www.amendiola.es/"),
    ("TDR Assessors", "Sant Boi de Llobregat", "info@tdrassessors.com", "https://tdrassessors.com/contacto/"),
    ("Assessoria Sant Boi S.L.", "Sant Boi de Llobregat", "josep@assessoriasantboi.cat", "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/assessoria-sant-boi_021273131_000000001.html"),
    ("Raurich Asesores", "Sant Boi de Llobregat", "info@raurich-asesores.com", "https://raurich-asesores.com/contacto/"),
    ("Balada Assessors", "Sant Boi de Llobregat", "olga@balada-assessors.com", "https://www.balada-assessors.com/es"),
    ("Gestoria Abella Gestió", "Sant Boi de Llobregat", None, "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/gestoria-abella-gestio_000581736_000000001.html"),
    ("Gabinet Assessor Ros", "Sant Boi de Llobregat", None, "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/gabinet-assessor-ros_172254500_000000001.html"),
    ("Bufet Assessor ADEC S.L.P.", "Sant Boi de Llobregat", None, "https://infonif.economia3.com/ficha-empresa/bufet-assessor-adec-slp"),
    ("Assessoria Integral de Sant Boi S.L.", "Sant Boi de Llobregat", None, "https://empresite.eleconomista.es/ASSESSORIA-INTEGRAL-SANT-BOI.html"),
    ("Vikmer Assessors S.L.P.", "Sant Boi de Llobregat", None, "https://www.gestorias.es/barcelona/sant-boi-de-llobregat/vikmer-assessors-19018"),
    ("Bgr Assessors", "Sant Boi de Llobregat", None, "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/bgr-assessors_159456037_000000001.html"),
    ("Gramalla XXI Assessors al seu Servei", "Sant Boi de Llobregat", None, "https://www.gramallaxxi.es/?lang=es"),
    ("Aba Serveis Empresarials S.L.", "Sant Boi de Llobregat", None, "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/aba-serveis-empresarials-s-l-_008982217_000000001.html"),
    ("Asesoria Coope S.L.", "Sant Boi de Llobregat", None, "https://www.asesoriacoope.es/es/"),
    ("I.M.S. Assessors Economics i Juridics S.L.", "Sant Boi de Llobregat", None, "https://www.paginasamarillas.es/a/asesoria-de-empresas/barcelona/sant-boi-de-llobregat/"),
    ("ASEMRECA S.L.U.", "Sant Boi de Llobregat", None, "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/asemreca-s-l-u-_196028658_000000002.html"),
    ("Arias Assessors (Sant Boi)", "Sant Boi de Llobregat", None, "https://www.carakter.org/arias-assessors"),
    ("La Asesoría Sant Boi", "Sant Boi de Llobregat", None, "https://laasesoriasantboi.com/contacto/"),

    ("Gestoria Simon & Torrent", "El Prat de Llobregat", "info@gestoriasimon.com", "https://www.gestoriasimon.com"),
    ("Fiscprat S.L.", "El Prat de Llobregat", "fiscal@fiscprat.es", "https://fiscprat.es/gestoria-en-el-prat-de-llobregat/"),
    ("Asesoria Carniago S.L.", "El Prat de Llobregat", "asesoria@carniago.com", "http://carniago.com/"),
    ("Piera Asesorías y Servicios S.L.", "El Prat de Llobregat", None, "https://www.asesoriapiera.com/"),
    ("Gabinete/Asesoría Delta", "El Prat de Llobregat", "gestion@asesoriadelta.com", "https://www.asesoriadelta.com/"),
    ("2A Assessors", "El Prat de Llobregat", "info@2aassessors.com", "https://www.2aassessors.com/"),
    ("Oficina de Gestión Centro S.L.P.", "El Prat de Llobregat", None, "https://www.qdq.com/oficina-de-gestion-centro-514265"),
    ("Forum Legal y Económico", "El Prat de Llobregat", "forum@forumlegalyeconomico.com", "https://www.forumlegalyeconomico.com/contacto/"),
    ("Tapia Gestoría Administrativa", "El Prat de Llobregat", None, "https://www.infoasesorias.es/tapia-gestoria-administrativa/"),
    ("K2 Assessors", "El Prat de Llobregat", None, "https://directorio.guia33.com/item/k2-assessors-el-prat/"),
    ("Grup 3 Assessors", "El Prat de Llobregat", "info@grup3assessors.com", "https://grup3assessors.com/contacto/"),
    ("Delaw Assessors S.L.", "El Prat de Llobregat", "delaw@delaw.es", "https://www.guia33.com/item/delaw-assessors-el-prat/"),
    ("MCC·GISE Asesoría", "El Prat de Llobregat", None, "https://asesoriamcc.es/"),
    ("Kosmos Asesores", "El Prat de Llobregat", "info@kosmosasesores.com", "https://kosmosasesores.com/"),
    ("ACM Asesores", "El Prat de Llobregat", "acm@acmsl.es", "https://www.asesoriaelprat.com/"),
    ("Asesoria Gestoria Integral de Pymes S.L.", "El Prat de Llobregat", None, "https://www.cylex.es/el-prat-de-llobregat/asesoria-gestoria-integral-de-pymes-s-l--12503696.html"),
    ("Gestoria de la Cámara-Jonama S.L.", "El Prat de Llobregat", None, "https://www.gestorias.es/barcelona/el-prat-de-llobregat/gestoria-de-la-camara-jonama-3764"),
    ("Fernández Rangel S.L.", "El Prat de Llobregat", None, "https://www.gestorias.es/barcelona/el-prat-de-llobregat"),
    ("P&G Consultores", "El Prat de Llobregat", None, "http://www.pygconsultores.net/"),
    ("Asesoria Ricart S.L.", "El Prat de Llobregat", None, "https://empresite.eleconomista.es/ASESORIA-RICART.html"),
    ("Asesoria Teerre S.L.", "El Prat de Llobregat", None, "https://www.einforma.com/informacion-empresa/asesoria-teerre"),
    ("Plaça Pau Casals S.C.P.", "El Prat de Llobregat", None, "https://www.paginasamarillas.es/f/el-prat-de-llobregat/placa-pau-casals-s-c-p-_200511459_000000002.html"),

    ("Gestoría BERNI", "Viladecans", "berni@grupfabrega.com", "https://www.gestoriabernifabrega.com"),
    ("GEMAP", "Viladecans", "gemap@gemap.es", "https://gemap.es/"),
    ("Ariza Gestoría", "Viladecans", "m-ariza@telefonica.net", "https://www.gestoria-ariza.es/"),
    ("Vistamar Gestión Empresarial", "Viladecans", "info@vistamar.cat", "http://vistamar.cat/"),
    ("Ingesdat 2008 S.L.", "Viladecans", "ingesdat@hotmail.es", "https://empresite.eleconomista.es/INGESDAT-2008.html"),
    ("Quatro Gestión", "Viladecans", "resteban@consultoriakoymark.com", "https://quatrogestion.es/"),
    ("Servicio Asesor Garantizado S.L.", "Viladecans", None, "https://www.paginasamarillas.es/f/viladecans/servicio-asesor-garantizado-s-l-_204687479_000000001.html"),
    ("DRI Assessoria", "Viladecans", None, "https://driassessoria.com/"),
    ("Asesoría Laboral Geslab S.L.", "Viladecans", None, "https://www.paginasamarillas.es/f/viladecans/asesoria-laboral-geslab-s-l-_021206677_000000001.html"),
    ("Baugar Gestora de Servicios", "Viladecans", None, "https://es.kompass.com/c/baugar-gestora-de-servicios/es1284495/"),
    ("Asesoría Areny", "Viladecans", None, "https://www.paginasamarillas.es/f/viladecans/asesoria-areny_008694648_000000001.html"),
    ("Fer Gestions 2009", "Viladecans", None, "https://empresite.eleconomista.es/FER-GESTIONS-2009.html"),
    ("Millán Gestió", "Viladecans", None, "https://www.infoasesorias.es/millan-gestio/"),
    ("Gómez i Carvacho Assessors", "Viladecans", None, "https://gomezcarvacho.com/nosotros/"),
    ("JJ & Assessors", "Viladecans", None, "https://assessoriajj.com/"),
    ("Arias Assessors (Viladecans)", "Viladecans", None, "https://www.carakter.org/arias-assessors"),
    ("Asesoría Empresarial Landa", "Viladecans", None, "https://aelanda.es/"),
    ("Gestram Associats 99", "Viladecans", None, "https://www.gestorias.es/barcelona/viladecans/gestram-associats-99-6966"),
    ("Sm Gestió", "Viladecans", None, "https://www.gestorias.es/barcelona/viladecans/sm-gestio-7263"),
    ("Paola Baquerizo Paladines", "Viladecans", None, "https://www.gestorias.es/barcelona/viladecans/2"),
    ("Rafael C. Gil", "Viladecans", None, "https://www.gestorias.es/barcelona/viladecans/2"),
    ("GIS, Grup Ip S.L.", "Viladecans", None, "https://www.gestorias.es/barcelona/viladecans/gis-grup-ip-16315"),

    ("Centre de Gestions Molins", "Molins de Rei", "aleix@centredegestions.com", "https://centredegestions.com/"),
    ("Molins Legal", "Molins de Rei", "info@molinslegal.com", "https://molinslegal.com/"),
    ("Sabaté i Matas Associats (Centre de Càlcul)", "Molins de Rei", "esteve@centredecalcul.net", "http://www.centredecalcul.net/"),
    ("E-D'ASS", "Molins de Rei", "e-dass@e-dass.com", "https://e-dass.com/"),
    ("Gestoría VP", "Molins de Rei", "info@gestoriavp.com", "https://gestoriavp.com/"),
    ("Gimeno Assessoria Jurídica (Gestoria Gimeno)", "Molins de Rei", "comptabilitats@gimeno.net", "https://www.gimeno.net/"),
    ("Asesoría Integral Gesticat Plus", "Molins de Rei", "info@gesticat.com", "https://gesticat.com/"),
    ("Alzueta & Saperas Assessors", "Molins de Rei", "recepcion@alzuetaysaperas.com", "https://alzuetaysaperas.com/contacto"),
    ("Gestions Empresarials Cabirol S.L.", "Molins de Rei", "info@cabirol.cat", "https://cabirolsl.com/contacto"),
    ("Tax Molins de Rei (Agesa)", "Molins de Rei", "asfilsa@agesa.es", "https://www.tax.es/es/oficinas/cataluna/tax-molins-de-rei.html"),
    ("Fàbrega Consultors", "Molins de Rei", None, "https://www.fabregaconsultors.com/"),
    ("Joan Tresserra Assessors S.L.", "Molins de Rei", None, "https://www.tresserra.cat/"),
    ("Bonafonte López S.C.P.", "Molins de Rei", None, "https://www.paginasamarillas.es/f/molins-de-rei/bonafonte-lopez-s-c-p-_180378630_000000001.html"),
    ("Solgemp Asesores", "Molins de Rei", None, "https://www.solgemp.com/"),
    ("Ferransa Asesores", "Molins de Rei", None, "https://ferransa.com/asesoria-fiscal-espana-molins-de-rei/"),
    ("Vernet Assessors i Associats S.L.", "Molins de Rei", None, "https://www.paginasamarillas.es/f/molins-de-rei/vernet-assessors-i-associats-s-l-_021411855_000000001.html"),
    ("Servei d'Assessorament Empresarial i Consultors S.L.", "Molins de Rei", None, "https://www.einforma.com/informacion-empresa/servei-assessorament-empresarial-consultors"),
    ("NINOAFIC S.A.", "Molins de Rei", None, "https://www.citiservi.es/barcelona/ninoafic-molins-de-rei__926211_940.html"),
    ("Grup Gestor Molins de Rei", "Molins de Rei", None, "https://www.cylex.es/molins-de-rei/grup-gestor-molins-de-rei-12661757.html"),
]

# ---------- 1) Extender cola_envios.xlsx ----------
wb_cola = load_workbook("cola_envios.xlsx")
ws_cola = wb_cola.active
headers = [c.value for c in ws_cola[1]]
col = {n: i for i, n in enumerate(headers)}

emails_en_cola = {str(ws_cola.cell(row=r, column=col["email"] + 1).value).strip().lower()
                  for r in range(2, ws_cola.max_row + 1) if ws_cola.cell(row=r, column=col["email"] + 1).value}

fechas_previstas = [ws_cola.cell(row=r, column=col["fecha_prevista"] + 1).value for r in range(2, ws_cola.max_row + 1)]
conteo_por_dia = {}
for f in fechas_previstas:
    if f:
        conteo_por_dia[f] = conteo_por_dia.get(f, 0) + 1
ultimo_dia = max(datetime.date.fromisoformat(f) for f in conteo_por_dia)

sin_email_nuevas = []  # (empresa, ciudad, fuente) para el maestro
nuevas_validas = []    # (nombre, empresa, ciudad, email, fuente) a añadir a la cola

for empresa, ciudad, email, fuente in NUEVAS:
    if not email:
        sin_email_nuevas.append((empresa, ciudad, fuente))
        continue
    email = email.strip().lower()
    if email in emails_en_cola:
        continue
    emails_en_cola.add(email)
    nuevas_validas.append((f"equip de {empresa}", empresa, ciudad, email, fuente))

dia_actual = ultimo_dia
for nombre, empresa, ciudad, email, fuente in nuevas_validas:
    while conteo_por_dia.get(dia_actual.isoformat(), 0) >= DAILY_CAP:
        dia_actual += datetime.timedelta(days=1)
    ws_cola.append([nombre, empresa, ciudad, "catalan", email, "", "", dia_actual.isoformat(), fuente])
    conteo_por_dia[dia_actual.isoformat()] = conteo_por_dia.get(dia_actual.isoformat(), 0) + 1

wb_cola.save("cola_envios.xlsx")
print(f"cola_envios.xlsx ampliada con {len(nuevas_validas)} gestorías nuevas (con email).")

# ---------- 2) Construir maestro_gestorias.xlsx ----------
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "maestro"
ws_out.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio",
               "fecha_prevista", "fuente_email", "comentario"])


def copiar_desde(path, excluir_pendientes_con_email=False):
    wb = load_workbook(path)
    ws = wb.active
    h = [c.value for c in ws[1]]
    c = {n: i for i, n in enumerate(h)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        estado = row[c["estado"]]
        email = row[c["email"]]
        if excluir_pendientes_con_email and estado != "enviado" and email:
            continue  # esta fila ya vive en cola_envios.xlsx, no duplicar
        estado_final = "enviado" if estado == "enviado" else ("sin email" if not email else "pendiente")
        ws_out.append([
            row[c["nombre"]], row[c["empresa"]], row[c["ciudad"]],
            "castellano" if "castellano" in path else "catalan",
            email or "", estado_final, row[c["fecha_envio"]] if estado == "enviado" else "",
            "", row[c["fuente_email"]] if "fuente_email" in c else "", "",
        ])


copiar_desde("contactos_catalan.xlsx")
copiar_desde("contactos_castellano.xlsx")
copiar_desde("contactos_catalan_ronda2.xlsx", excluir_pendientes_con_email=True)

# cola_envios.xlsx: todo pendiente, con fecha_prevista
wb_cola = load_workbook("cola_envios.xlsx")
ws_cola = wb_cola.active
h = [c.value for c in ws_cola[1]]
c = {n: i for i, n in enumerate(h)}
for row in ws_cola.iter_rows(min_row=2, values_only=True):
    estado = "enviado" if row[c["estado"]] == "enviado" else "pendiente"
    ws_out.append([
        row[c["nombre"]], row[c["empresa"]], row[c["ciudad"]], row[c["idioma"]],
        row[c["email"]], estado, row[c["fecha_envio"]] or "", row[c["fecha_prevista"]] or "",
        row[c["fuente_email"]] or "", "",
    ])

# gestorías de las 6 ciudades nuevas sin email encontrado
for empresa, ciudad, fuente in sin_email_nuevas:
    ws_out.append([f"equip de {empresa}", empresa, ciudad, "catalan", "", "sin email", "", "", fuente, ""])

wb_out.save("maestro_gestorias.xlsx")

total = ws_out.max_row - 1
print(f"maestro_gestorias.xlsx creado con {total} filas.")
