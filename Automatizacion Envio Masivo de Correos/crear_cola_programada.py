"""Construye cola_envios.xlsx: cola combinada de gestorías pendientes (las que quedaron
sin enviar de la ronda 2 por el límite de Hostinger + las 6 ciudades nuevas encontradas
julio 2026), con una columna fecha_prevista que reparte los envíos a razón de
DAILY_CAP por día a partir de mañana (hoy ya se superó el límite diario)."""
import datetime
from openpyxl import Workbook, load_workbook

DAILY_CAP = 105
HOY = datetime.date(2026, 7, 13)
PRIMER_DIA = HOY + datetime.timedelta(days=1)

# emails ya usados (para no duplicar destinatarios entre ciudades distintas)
YA_USADOS = {"assessoria@escabiassessors.com"}

# (empresa, ciudad, email_o_None, fuente)
NUEVAS = [
    ("Vilanova Gestió S.L.", "Vilanova i la Geltrú", "gestoria@vilanovagestio.com", "https://www.vilanovagestio.com/"),
    ("Gestal Assessors", "Vilanova i la Geltrú", "info@gestalassessors.com", "https://gestalassessors.com/"),
    ("Vidal Piqué Assessors", "Vilanova i la Geltrú", "vidal@vidalpiqueassessors.cat", "https://www.vidalpiqueassessors.cat/"),
    ("Grup de Gestió Fiscal (GGF)", "Vilanova i la Geltrú", "ggf@grupgestiofiscal.com", "https://grupgestiofiscal.com/"),
    ("Premià Assessors S.C.C.L.", "Vilanova i la Geltrú", "info@premiavng.com", "https://premiavng.com/"),
    ("Gestoría Hernández", "Vilanova i la Geltrú", "hernandez@gestors.net", "https://www.gestorshernandez.es/es/asesoria-de-empresas-vilanova-i-la-geltru"),
    ("9 Assessors", "Vilanova i la Geltrú", "9@9assessors.com", "https://www.9assessors.com/es/asesores-en-vilanova/"),
    ("Ceteb", "Vilanova i la Geltrú", "ceteb@ceteb.com", "https://ceteb.com/"),
    ("Gespas Gestoria", "Vilanova i la Geltrú", "info@gespasgestoria.com", "https://gespasgestoria.com/contacto/"),
    ("Martinez Assessors de Vilanova S.L.", "Vilanova i la Geltrú", "info@martinezassessors.com", "https://martinezassessors.com/"),
    ("Solfico", "Vilanova i la Geltrú", "solfico@solfico.es", "https://solfico.es/asesoria-en-vilanova-i-la-geltru/"),
    ("Inval Gestoria", "Vilanova i la Geltrú", "info@invalgestoria.com", "https://invalgestoria.com/"),
    ("Gestiona (Gestions Solucions)", "Vilanova i la Geltrú", "gestiona@gestionasolucions.cat", "https://www.paginasamarillas.es/f/vilanova-i-la-geltru/gestiona_221854276_000000001.html"),
    ("Bufet Grup Quatre", "Vilanova i la Geltrú", "grup4@grupquatre.es", "https://grupquatre.es/"),
    ("A&A Assessors S.C.P.", "Vilanova i la Geltrú", "ayases@ayases.es", "https://www.ayases.com/es/contacto/"),
    ("Grup Integra Consulting", "Vilanova i la Geltrú", "info@grupintegraconsulting.com", "https://grupintegraconsulting.com/contacto/"),
    ("Ardebol Assessors, S.L.", "Vilanova i la Geltrú", "acantos@ardebolassessors.cat", "https://www.gestorias.es/barcelona/vilanova-i-la-geltru/ardebol-assessors-17917"),
    ("Gestoria Martorell", "Vilanova i la Geltrú", "info@gestoriamartorell.com", "http://www.gestoriamartorell.com/contacto"),

    ("Escabia Assessors S.L.P.", "Cerdanyola del Vallès", "assessoria@escabiassessors.com", "https://www.gestorias.es/barcelona/cerdanyola-del-valles/escabia-assessors-9334"),
    ("Gestoria Vico", "Cerdanyola del Vallès", "info@gestoriacerdanyola.com", "https://gestoriacerdanyola.com/"),
    ("Grup Sisquella (La Gestora Cerdanyola)", "Cerdanyola del Vallès", "info@grupsisquella.com", "https://grupsisquella.com/"),
    ("Gestoría Ángel González (SG Assessors)", "Cerdanyola del Vallès", "info@gaag.es", "https://gaag.es/"),
    ("GrupDos", "Cerdanyola del Vallès", "grupdos@grupdos.com", "https://www.grupdos.com/es/contacte.html"),
    ("Gestió Integral Assessors Barcelona", "Cerdanyola del Vallès", "info@gestoriasbarcelona.com", "https://gestoriasbarcelona.com/"),
    ("Rosa Gestión S.A.", "Cerdanyola del Vallès", "administracion@a-csn.com", "https://empresite.eleconomista.es/ROSA-GESTION.html"),
    ("Assessoria Cerdanyola S.L.", "Cerdanyola del Vallès", "info@assessoriacerdanyola.com", "https://empresite.eleconomista.es/ASSESSORIA-CERDANYOLA.html"),
    ("Gestoria Albadalejo", "Cerdanyola del Vallès", "albadalejo@gestors.net", "https://asesoriacerca.es/info/25592-gestoria-albadalejo"),

    ("Serveis Administratius Simon S.L.", "Sant Cugat del Vallès", "simon@serveissimon.com", "https://empresite.eleconomista.es/SERVEIS-ADMINISTRATIUS-SIMON.html"),
    ("Gap Assessors", "Sant Cugat del Vallès", "info@gapassessors.com", "http://www.gapassessors.com/contacte/"),
    ("Sant Cugat Consulting", "Sant Cugat del Vallès", "scc@santcugatconsulting.com", "https://www.santcugatconsulting.es/contacto/"),
    ("Tràmit Sant Cugat S.L.", "Sant Cugat del Vallès", "administracio@tramitserveis.cat", "https://tramitserveis.cat/es/contacto/"),
    ("G.A.C. Grupo de Asesoramiento y Consulting S.L.", "Sant Cugat del Vallès", "gacgrup@gacgrup.com", "https://www.gacgrup.com/contacto/"),
    ("Edit Economistes Consultors S.L.P.", "Sant Cugat del Vallès", "edit@editconsultores.com", "https://empresite.eleconomista.es/EDIT-ECONOMISTES-CONSULTORS-SLP.html"),
    ("Plus Gestió", "Sant Cugat del Vallès", "info@plusgestio.com", "https://plusgestio.com/"),
    ("Asenta Consulting S.L.", "Sant Cugat del Vallès", "asenta@asenta.es", "https://www.asenta.es/en/contact/"),
    ("Indeed Asesores S.L.", "Sant Cugat del Vallès", "info@indeedasesores.com", "https://empresite.eleconomista.es/INDEED-ASESORES.html"),
    ("Gestoria Administrativa Jordi Juncosas", "Sant Cugat del Vallès", "info@gestoriajuncosas.com", "https://www.gestoria-santcugat.com/precio-gestoria-sant-cugat-contacto.html"),
    ("Administración de Empresas Sant Cugat S.L.", "Sant Cugat del Vallès", "pablo@adesantcugat.es", "https://empresite.eleconomista.es/ADMINISTRACION-EMPRESAS-SANT-CUGAT.html"),

    ("Gestoria Capmany", "Rubí", "josep@gestoriacapmany.com", "https://www.gestoriacapmany.es/contacto"),
    ("Rubigest Assessors", "Rubí", "rubigest@rubigest.com", "https://www.rubigest.com/"),
    ("Gestoria Julià", "Rubí", "gestoriajulia@gmail.com", "https://gestoriajulia.com/contacto/"),
    ("Rubí Asesores", "Rubí", "info@rubiasesores.com", "https://rubiasesores.com/"),
    ("Consultoria Moreno Martínez", "Rubí", "antonio@consultoriamorenomartinez.com", "https://consultoriamorenomartinez.com/"),
    ("Asesores GS (González Segura Asesores)", "Rubí", "info@asesoresgs.com", "https://asesoresgs.com/"),
    ("Gabinet Rovira", "Rubí", "rovira@gabinet-rovira.com", "https://gabinet-rovira.com/gestoria-en-rubi/"),
    ("CET 4 Gestió Empresarial S.L.", "Rubí", "cet4@cet4.es", "https://empresite.eleconomista.es/CET-4-GESTIO-EMPRESARIAL.html"),
    ("Busbac Serveis (Busbac+)", "Rubí", "hola@busbac.com", "https://busbac.com/contacts-es/"),
    ("Gestoria Ridao (Ridao Asesores)", "Rubí", "ridao@gesridao.com", "https://www.gesridao.com/contacte/contacte.php?lg=es"),
    ("Escabia Assessors (Rubí)", "Rubí", "assessoria@escabiassessors.com", "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/rubi/"),

    ("Gestoria Vilafranca (Jané Associats)", "Vilafranca del Penedès", "info@gestoriavilafranca.com", "https://gestoriavilafranca.com/"),
    ("Cyp Assessors", "Vilafranca del Penedès", "info@cypassessors.cat", "https://www.cypassessors.cat/"),
    ("De Cara Assessors", "Vilafranca del Penedès", "info@decaraassessors.com", "https://www.gestdecara.com/"),
    ("Planas i Associats (Gestoria Planas)", "Vilafranca del Penedès", "info@gestoriaplanas.com", "http://www.gestoriaplanas.com/"),
    ("Gestoría Olivella", "Vilafranca del Penedès", "info@gestoriaolivella.com", "https://gestoriaolivella.com/"),
    ("Gestoria Solé", "Vilafranca del Penedès", "acobo@gestoriasole.com", "https://gestoriasole.com/es/contacto/"),
    ("Consulting Penedès S.L.", "Vilafranca del Penedès", "admin@consultingpenedes.cat", "https://www.consultingpenedes.com/"),
    ("Gestoria Freixedas", "Vilafranca del Penedès", "freixedas@gestors.net", "https://gestoriafreixedas.es/"),
    ("Gestoria Estalella (Serestal SLP)", "Vilafranca del Penedès", "hola@serestal.cat", "https://mundoabogado.com/abogado-en-vilafranca-del-penedes-serestal-slp-gestoria-estalella-18705"),
    ("Baimpost S.L.", "Vilafranca del Penedès", "csdbaimpost@msn.com", "https://www.baimpost.com/ubicacion.html"),
    ("Osorio Consulting S.L.P.", "Vilafranca del Penedès", "gestoria@osorioconsulting.net", "https://www.osorioconsulting.net/"),
    ("Solucions Empresarials Penedès", "Vilafranca del Penedès", "info@solucionsempresarialspenedes.com", "https://solucionsempresarialspenedes.com/"),
    ("1+1 Consultoria de Gestió", "Vilafranca del Penedès", "hola@1mes1gestio.com", "https://1mes1gestio.com/es/contacto/"),
    ("Assessoria Calvet", "Vilafranca del Penedès", "contacte@assessoriacalvet.cat", "http://assessoriacalvet.cat/"),

    ("CINC Centre de Negocis", "Figueres", "administracio@cinc.es", "https://www.cinc.com/ca/contacte/"),
    ("Grup Simon Global", "Figueres", "info@grupsimon.com", "https://grupsimonglobal.com/contacte/figueres/"),
    ("Gestoria Santiago Fernández Carbó", "Figueres", "info@gestoriafernandez.com", "http://www.gestoriafernandez.com/es/contactar.html"),
    ("Aster (Asterfiscal)", "Figueres", "asterscfigueres@yahoo.es", "https://asterfiscal.es/"),
    ("Gestoria Mensión", "Figueres", "gestoria@mension.net", "http://www.mension.net/"),
    ("Mediterrània de Serveis", "Figueres", "info@gestoriamediterrania.com", "http://www.gestoriamediterrania.com/"),
    ("Assessoria Prats", "Figueres", "info@pratsassessoria.com", "https://pratsassessoria.com/es/"),
    ("Gestoria Bartolomé", "Figueres", "info@gestoriabartolome.cat", "https://gestoriabartolome.com/"),
    ("Assessoria Roig-Casamitjana", "Figueres", "roigcorreduria@hotmail.com", "https://www.paginasamarillas.es/f/figueres/assessoria-roig-casamitjana_018920884_000000001.html"),
    ("LaIA Consulting", "Figueres", "info@laiaconsulting.com", "https://www.laiaconsulting.com/contacte/"),
    ("Vivogest - Dabau", "Figueres", "info@vivogest.cat", "https://vivogest.cat/"),
    ("Costa Clotas Comptabilitat i Gestió", "Figueres", "costaclotas@economistes.com", "http://www.costaclotas.com/"),
    ("Avincla", "Figueres", "barcelona@avincla.com", "https://www.avincla.com/"),
    ("Assessoria Teixidor", "Figueres", "atclient@assessoria-teixidor.com", "http://www.assessoria-teixidor.com/"),
    ("Quatredelnord Assessors d'Empresa", "Figueres", "qdn@qdn.cat", "http://www.qdn.cat/"),
    ("Reisu Assessors", "Figueres", "reisu@reisuassessors.com", "https://reisuassessors.com/"),
    ("RM Assessors Figueres", "Figueres", "rupertsanllehi@rm-assessors.cat", "http://www.rm-assessors.cat/"),
    ("Tax Figueres Economistes i Advocats", "Figueres", "central@tax.es", "https://www.tax.es/en/catalonia/figueres-tax-consulting/company-advice-legal-fiscal-employment"),
    ("Assessoria Barceló", "Figueres", "info@assessoriabarcelo.com", "https://www.assessoriabarcelo.com/spa/"),
    ("AVL Assessors", "Figueres", "info@avlassessors.com", "https://avlassessors.com/?lang=es"),
    ("Delta Grup Assessors", "Figueres", "gerencia@deltaglobal.es", "https://www.gestorias.es/girona/figueres/delta-grup-assessors-15363"),
]

wb_r2 = load_workbook("contactos_catalan_ronda2.xlsx")
ws_r2 = wb_r2.active
headers_r2 = [c.value for c in ws_r2[1]]
col_r2 = {n: i for i, n in enumerate(headers_r2)}

filas = []  # (nombre, empresa, ciudad, idioma, email, fuente)
vistos = set()

for row in ws_r2.iter_rows(min_row=2, values_only=True):
    if row[col_r2["estado"]] != "enviado" and row[col_r2["email"]]:
        email = row[col_r2["email"]].strip()
        if email in vistos or email in YA_USADOS:
            continue
        vistos.add(email)
        filas.append((row[col_r2["nombre"]], row[col_r2["empresa"]], row[col_r2["ciudad"]], "catalan", email, row[col_r2["fuente_email"]]))

for empresa, ciudad, email, fuente in NUEVAS:
    if not email:
        continue
    if email in vistos or email in YA_USADOS:
        continue
    vistos.add(email)
    filas.append((f"equip de {empresa}", empresa, ciudad, "catalan", email, fuente))

wb = Workbook()
ws = wb.active
ws.title = "cola"
ws.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio", "fecha_prevista", "fuente_email"])

for i, (nombre, empresa, ciudad, idioma, email, fuente) in enumerate(filas):
    dia = PRIMER_DIA + datetime.timedelta(days=i // DAILY_CAP)
    ws.append([nombre, empresa, ciudad, idioma, email, "", "", dia.isoformat(), fuente])

wb.save("cola_envios.xlsx")

dias = (len(filas) - 1) // DAILY_CAP + 1
print(f"Cola creada: {len(filas)} gestorías pendientes, repartidas en {dias} día(s) a {DAILY_CAP}/día, empezando el {PRIMER_DIA.isoformat()}.")
