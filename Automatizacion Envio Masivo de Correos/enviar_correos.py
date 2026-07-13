"""Envío masivo de correos personalizados a partir de una plantilla HTML y un Excel de contactos.

El Excel se usa como fuente y como registro: cada fila enviada se marca en la
columna 'estado' (y 'fecha_envio'), así que se puede cortar el proceso y
reanudarlo más tarde sin repetir correos.
"""
import argparse
import os
import re
import smtplib
import ssl
import sys
import time
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from dotenv import load_dotenv
from openpyxl import load_workbook

PLACEHOLDER_RE = re.compile(r"\{\{(\w+)\}\}")


def render(text, data):
    def replace(match):
        value = data.get(match.group(1))
        return "" if value is None else str(value)
    return PLACEHOLDER_RE.sub(replace, text)


def strip_html(html):
    return re.sub(r"<[^>]+>", "", html)


def connect_smtp(host, port, user, password):
    if port == 465:
        server = smtplib.SMTP_SSL(host, port, context=ssl.create_default_context())
    else:
        server = smtplib.SMTP(host, port)
        server.starttls(context=ssl.create_default_context())
    server.login(user, password)
    return server


def enviar_email(server, from_addr, from_name, to_addr, subject, html_body):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"{from_name} <{from_addr}>"
    msg["To"] = to_addr
    msg.attach(MIMEText(strip_html(html_body), "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    server.sendmail(from_addr, [to_addr], msg.as_string())


def cargar_config():
    load_dotenv()
    faltantes = [k for k in ("SMTP_HOST", "SMTP_USER", "SMTP_PASS") if not os.environ.get(k)]
    if faltantes:
        sys.exit(f"Faltan variables en .env: {', '.join(faltantes)} (copia .env.example a .env y rellénalo)")
    return {
        "host": os.environ["SMTP_HOST"],
        "port": int(os.environ.get("SMTP_PORT", 465)),
        "user": os.environ["SMTP_USER"],
        "password": os.environ["SMTP_PASS"],
        "from_name": os.environ.get("FROM_NAME", os.environ["SMTP_USER"]),
        "delay": float(os.environ.get("DELAY_SECONDS", 8)),
    }


def main():
    parser = argparse.ArgumentParser(description="Envío masivo de correos personalizados")
    parser.add_argument("--excel", required=True, help="Ruta al Excel de contactos (.xlsx)")
    parser.add_argument("--template", required=True, help="Ruta a la plantilla HTML")
    parser.add_argument("--asunto", required=True, help="Asunto del correo, admite {{marcadores}}")
    parser.add_argument("--hoja", default=None, help="Nombre de la hoja (por defecto: la activa)")
    parser.add_argument("--col-email", default="email", help="Columna con el email (por defecto: email)")
    parser.add_argument("--col-estado", default="estado", help="Columna para marcar enviado/error")
    parser.add_argument("--col-fecha", default="fecha_envio", help="Columna con la fecha de envío")
    parser.add_argument("--limite", type=int, default=None, help="Máximo de correos a enviar en esta ejecución")
    parser.add_argument("--dry-run", action="store_true", help="Simula sin enviar correos ni tocar el Excel")
    parser.add_argument("--prueba-a", default=None, help="Envía solo un correo de prueba (fila 2) a esta dirección")
    args = parser.parse_args()

    config = cargar_config()

    with open(args.template, encoding="utf-8") as f:
        template_html = f.read()

    wb = load_workbook(args.excel)
    ws = wb[args.hoja] if args.hoja else wb.active

    headers = [c.value for c in ws[1]]
    col_index = {name: i + 1 for i, name in enumerate(headers) if name}

    def ensure_column(name):
        if name not in col_index:
            nueva = ws.max_column + 1
            ws.cell(row=1, column=nueva, value=name)
            col_index[name] = nueva
            headers.append(name)
        return col_index[name]

    if args.col_email not in col_index:
        sys.exit(f"No encuentro la columna '{args.col_email}' en la primera fila de {args.excel}.")
    estado_col = ensure_column(args.col_estado)
    fecha_col = ensure_column(args.col_fecha)

    def fila_a_datos(row):
        return {headers[i]: ws.cell(row=row, column=i + 1).value for i in range(len(headers)) if headers[i]}

    server = connect_smtp(**{k: config[k] for k in ("host", "port", "user", "password")}) if not args.dry_run else None

    try:
        if args.prueba_a:
            data = fila_a_datos(2)
            subject = render(args.asunto, data)
            body = render(template_html, data)
            if args.dry_run:
                print(f"[DRY-RUN] Asunto: {subject}\n\n{body}")
            else:
                enviar_email(server, config["user"], config["from_name"], args.prueba_a, subject, body)
                print(f"Correo de prueba enviado a {args.prueba_a}")
            return

        enviados = saltados = errores = 0
        for row in range(2, ws.max_row + 1):
            if args.limite and enviados >= args.limite:
                break

            if ws.cell(row=row, column=estado_col).value == "enviado":
                saltados += 1
                continue

            data = fila_a_datos(row)
            to_addr = data.get(args.col_email)
            if not to_addr:
                continue

            subject = render(args.asunto, data)
            body = render(template_html, data)

            if args.dry_run:
                print(f"[DRY-RUN] -> {to_addr} | {subject}")
                enviados += 1
                continue

            try:
                enviar_email(server, config["user"], config["from_name"], to_addr, subject, body)
                ws.cell(row=row, column=estado_col, value="enviado")
                ws.cell(row=row, column=fecha_col, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                enviados += 1
                print(f"Enviado a {to_addr}")
            except Exception as e:
                ws.cell(row=row, column=estado_col, value=f"error: {e}")
                errores += 1
                print(f"Error enviando a {to_addr}: {e}")

            wb.save(args.excel)
            time.sleep(config["delay"])

        print(f"\nResumen: {enviados} enviados, {saltados} ya estaban enviados, {errores} errores.")
    finally:
        if server:
            server.quit()
        if not args.dry_run:
            wb.save(args.excel)


if __name__ == "__main__":
    main()
