"""Separa contactos_gestorias.xlsx en dos archivos según idioma del envío:
catalán (Vic, Sabadell, Terrassa, Igualada) y castellano (el resto, ej. Zaragoza, Asturias)."""
from openpyxl import Workbook, load_workbook

CIUDADES_CATALAN = {"Vic", "Sabadell", "Terrassa", "Igualada"}

wb = load_workbook("contactos_gestorias.xlsx")
ws = wb.active
headers = [c.value for c in ws[1]]
ciudad_col = headers.index("ciudad")

wb_cat = Workbook()
ws_cat = wb_cat.active
ws_cat.title = "gestorias"
ws_cat.append(headers)

wb_es = Workbook()
ws_es = wb_es.active
ws_es.title = "gestorias"
ws_es.append(headers)

n_cat = n_es = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[ciudad_col] in CIUDADES_CATALAN:
        ws_cat.append(row)
        n_cat += 1
    else:
        ws_es.append(row)
        n_es += 1

wb_cat.save("contactos_catalan.xlsx")
wb_es.save("contactos_castellano.xlsx")
print(f"contactos_catalan.xlsx: {n_cat} filas")
print(f"contactos_castellano.xlsx: {n_es} filas")
