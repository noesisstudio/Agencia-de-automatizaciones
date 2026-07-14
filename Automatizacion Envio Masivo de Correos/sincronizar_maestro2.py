"""Segunda pasada de sincronización: los 5 emails que faltaban (Gestoria Font,
Campdepadros Gestoria, Corporació Raif S.L., Asde Assessors 2002, Vives i Roig S.L.)
viven en contactos_catalan_ronda2.xlsx, no en cola_envios.xlsx. Los actualiza ahí
y los añade a cola_envios.xlsx para que puedan enviarse, reparte fecha_prevista y
reconstruye maestro_gestorias.xlsx."""
import datetime
from openpyxl import Workbook, load_workbook

DAILY_CAP = 95
PRIMER_DIA = datetime.date(2026, 7, 14)

wb_m = load_workbook("maestro_gestorias.xlsx")
ws_m = wb_m.active
hm = [c.value for c in ws_m[1]]
cm = {n: i for i, n in enumerate(hm)}

pendientes_sync = {}
for row in ws_m.iter_rows(min_row=2, values_only=True):
    if row[cm["estado"]] == "sin email" and row[cm["email"]] and "@" in str(row[cm["email"]]):
        pendientes_sync[row[cm["empresa"]]] = row[cm["email"]].strip()

print("Pendientes de sincronizar:", pendientes_sync)

# 1) actualizar contactos_catalan_ronda2.xlsx donde corresponda
wb_r2 = load_workbook("contactos_catalan_ronda2.xlsx")
ws_r2 = wb_r2.active
hr2 = [c.value for c in ws_r2[1]]
cr2 = {n: i + 1 for i, n in enumerate(hr2)}

filas_nuevas_para_cola = []  # (nombre, empresa, ciudad, email, fuente)
for r in range(2, ws_r2.max_row + 1):
    empresa = ws_r2.cell(row=r, column=cr2["empresa"]).value
    if empresa in pendientes_sync and not ws_r2.cell(row=r, column=cr2["email"]).value:
        email = pendientes_sync.pop(empresa)
        ws_r2.cell(row=r, column=cr2["email"], value=email)
        ciudad = ws_r2.cell(row=r, column=cr2["ciudad"]).value
        nombre = ws_r2.cell(row=r, column=cr2["nombre"]).value
        fuente = ws_r2.cell(row=r, column=cr2["fuente_email"]).value if "fuente_email" in cr2 else ""
        filas_nuevas_para_cola.append((nombre, empresa, ciudad, email, fuente))

wb_r2.save("contactos_catalan_ronda2.xlsx")
print(f"contactos_catalan_ronda2.xlsx actualizado, {len(filas_nuevas_para_cola)} filas listas para pasar a la cola.")

# 2) resto (si queda algo) intentar en cola_envios.xlsx (por si acaso)
wb = load_workbook("cola_envios.xlsx")
ws = wb.active
h = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(h)}

for r in range(2, ws.max_row + 1):
    empresa = ws.cell(row=r, column=col["empresa"]).value
    if empresa in pendientes_sync and not ws.cell(row=r, column=col["email"]).value:
        email = pendientes_sync.pop(empresa)
        ws.cell(row=r, column=col["email"], value=email)
        ws.cell(row=r, column=col["estado"], value="")

if pendientes_sync:
    print("AVISO: no encontrados en ningún archivo:", pendientes_sync)

# 3) añadir las filas movidas desde ronda2 a cola_envios.xlsx
for nombre, empresa, ciudad, email, fuente in filas_nuevas_para_cola:
    ws.append([nombre, empresa, ciudad, "catalan", email, "", "", "", fuente])

# 4) repartir fecha_prevista (mismo orden de fichero, para no mover lo ya programado hoy)
pendientes_rows = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(row=r, column=col["estado"]).value != "enviado"
                   and ws.cell(row=r, column=col["email"]).value]

for i, r in enumerate(pendientes_rows):
    dia = PRIMER_DIA + datetime.timedelta(days=i // DAILY_CAP)
    ws.cell(row=r, column=col["fecha_prevista"], value=dia.isoformat())

wb.save("cola_envios.xlsx")

# --- reconstruir maestro_gestorias.xlsx ---
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "maestro"
ws_out.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio",
               "fecha_prevista", "fuente_email", "comentario"])


def copiar_desde(path, excluir_pendientes_con_email=False):
    wbx = load_workbook(path)
    wsx = wbx.active
    hx = [c.value for c in wsx[1]]
    cx = {n: i for i, n in enumerate(hx)}
    for row in wsx.iter_rows(min_row=2, values_only=True):
        estado = row[cx["estado"]]
        email = row[cx["email"]]
        if excluir_pendientes_con_email and estado != "enviado" and email:
            continue
        estado_final = "enviado" if estado == "enviado" else ("sin email" if not email else "pendiente")
        ws_out.append([
            row[cx["nombre"]], row[cx["empresa"]], row[cx["ciudad"]],
            "castellano" if "castellano" in path else "catalan",
            email or "", estado_final, row[cx["fecha_envio"]] if estado == "enviado" else "",
            "", row[cx["fuente_email"]] if "fuente_email" in cx else "", "",
        ])


copiar_desde("contactos_catalan.xlsx")
copiar_desde("contactos_castellano.xlsx")
copiar_desde("contactos_catalan_ronda2.xlsx", excluir_pendientes_con_email=True)

wb_cola = load_workbook("cola_envios.xlsx")
ws_cola = wb_cola.active
hc = [c.value for c in ws_cola[1]]
cc = {n: i for i, n in enumerate(hc)}
for row in ws_cola.iter_rows(min_row=2, values_only=True):
    if row[cc["estado"]] == "enviado":
        estado_final = "enviado"
    elif not row[cc["email"]]:
        estado_final = "sin email"
    else:
        estado_final = "pendiente"
    ws_out.append([
        row[cc["nombre"]], row[cc["empresa"]], row[cc["ciudad"]], row[cc["idioma"]],
        row[cc["email"]] or "", estado_final, row[cc["fecha_envio"]] or "", row[cc["fecha_prevista"]] or "",
        row[cc["fuente_email"]] or "", "",
    ])

wb_out.save("maestro_gestorias.xlsx")
print(f"maestro_gestorias.xlsx reconstruido con {ws_out.max_row - 1} filas.")
