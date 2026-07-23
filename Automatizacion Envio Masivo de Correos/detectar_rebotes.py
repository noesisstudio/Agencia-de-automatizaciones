"""Detecta rebotes (bounces) en el buzón vía IMAP y los marca en maestro_gestorias.xlsx.

Solo lee mensajes de tipo 'delivery status' (Mail Delivery System / postmaster);
no toca ni lee el resto del correo. Marca cada dirección rebotada como
'rebotado' en la columna estado del maestro (sin tocar la fecha de envío,
para conservar el histórico de cuándo se intentó).

Uso: .venv\\Scripts\\python.exe detectar_rebotes.py [--dias 3]
"""
import argparse
import email
import imaplib
import os
import re
from datetime import date, timedelta

from dotenv import load_dotenv
from openpyxl import load_workbook

IMAP_HOST = "imap.hostinger.com"
IMAP_PORT = 993

RE_EMAIL = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
ASUNTOS_REBOTE = ("undelivered", "undeliverable", "returned to sender", "delivery status", "failure")


def extraer_rebotados(msg):
    """Extrae direcciones fallidas de un mensaje de rebote (cabecera DSN o cuerpo)."""
    rebotados = set()
    for parte in msg.walk():
        ctype = parte.get_content_type()
        if ctype == "message/delivery-status":
            payload = parte.get_payload()
            partes = payload if isinstance(payload, list) else [payload]
            for p in partes:
                texto = p.as_string() if hasattr(p, "as_string") else str(p)
                for m in re.finditer(r"(?:Final|Original)-Recipient:.*?;\s*([^\s;]+@[^\s;]+)", texto, re.I):
                    rebotados.add(m.group(1).strip().lower().rstrip(">"))
        elif ctype in ("text/plain", "text/html") and not rebotados:
            try:
                texto = parte.get_payload(decode=True).decode(parte.get_content_charset() or "utf-8", "replace")
            except Exception:
                continue
            for m in re.finditer(r"<?([\w.+-]+@[\w-]+\.[\w.-]+)>?:?\s*(?:host|mailbox|user|address|550|554)", texto, re.I):
                candidato = m.group(1).lower()
                if not candidato.endswith("bynoesis.com"):
                    rebotados.add(candidato)
    return rebotados


def main():
    parser = argparse.ArgumentParser(description="Detecta rebotes por IMAP y actualiza el maestro")
    parser.add_argument("--dias", type=int, default=3, help="Buscar rebotes de los últimos N días")
    parser.add_argument("--maestro", default="maestro_gestorias.xlsx")
    args = parser.parse_args()

    load_dotenv()
    user = os.environ["SMTP_USER"]
    password = os.environ["SMTP_PASS"]

    imap = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    imap.login(user, password)
    imap.select("INBOX", readonly=True)

    desde = (date.today() - timedelta(days=args.dias)).strftime("%d-%b-%Y")
    _, datos = imap.search(None, f'(SINCE "{desde}")')
    ids = datos[0].split()

    rebotados = set()
    for mid in ids:
        _, cabecera = imap.fetch(mid, "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT)])")
        cab = (cabecera[0][1] or b"").decode("utf-8", "replace").lower()
        es_rebote = ("mailer-daemon" in cab or "postmaster" in cab or "mail delivery" in cab
                     or any(k in cab for k in ASUNTOS_REBOTE))
        if not es_rebote:
            continue
        _, cuerpo = imap.fetch(mid, "(BODY.PEEK[])")
        msg = email.message_from_bytes(cuerpo[0][1])
        rebotados |= extraer_rebotados(msg)

    imap.logout()

    if not rebotados:
        print("No se han encontrado rebotes en el período.")
        return

    print(f"Direcciones rebotadas detectadas ({len(rebotados)}):")
    for r in sorted(rebotados):
        print(f"  - {r}")

    wb = load_workbook(args.maestro)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(headers)}

    marcados = 0
    for r in range(2, ws.max_row + 1):
        email_fila = ws.cell(row=r, column=col["email"]).value
        if email_fila and email_fila.strip().lower() in rebotados:
            if ws.cell(row=r, column=col["estado"]).value != "rebotado":
                ws.cell(row=r, column=col["estado"], value="rebotado")
                marcados += 1

    wb.save(args.maestro)
    print(f"\n{marcados} filas marcadas como 'rebotado' en {args.maestro}.")


if __name__ == "__main__":
    main()
