"""Añade a cola_envios.xlsx la ronda 8 (Valls, El Vendrell, Cambrils — búsqueda
20/07/2026), deduplicando contra todos los emails ya usados. Después ejecutar:
reconstruir_maestro.py"""
from datetime import date

from openpyxl import load_workbook

FECHA_HOY = date.today().isoformat()

RONDA8 = [
    ("Gabinet Assessor Alt Camp", "Valls", "gabassac@gabassac.cat", "https://www.topasesorias.com/info/gabinet-assessor-alt-camp-valls"),
    ("Lopez i Valles Associats S.L.", "Valls", None, "https://empresite.eleconomista.es/LOPEZ-VALLES-ASSESSORS.html"),
    ("Assessoria Gescova", "Valls", "gescova@gescova.cat", "https://www.gescova.cat/es/contacto/"),
    ("Gestions Palau", "Valls", "info@gestionspalau.es", "https://gestionspalau.es/"),
    ("Estrangeria - Isabel March Altes", "Valls", None, "https://www.paginasamarillas.es/search/asesorias-fiscales/all-ma/tarragona/all-is/valls/all-ba/all-pu/all-nc/1"),
    ("LaborGest Valls", "Valls", "info@laborgestvalls.com", "https://laborgestvalls.com/contacta-amb-laborgest/"),
    ("Gestoria Casañas", "Valls", "info@casanasassessors.com", "https://www.serveisactius.cat/es/altcamp/asesoria-valls"),
    ("Coordinadora de Gestión de Ingresos S.A.", "Valls", None, "https://www.gestorias.es/tarragona/valls/coordinadora-de-gestion-de-ingresos-3474"),
    ("Gestoria Escoda S.L.", "Valls", None, "https://www.paginasamarillas.es/f/valls/gestoria-escoda-s-l-_180787517_000000001.html"),
    ("BNA Consultors S.L.", "Valls", "comunicacio@bna.cat", "https://www.bna.cat/"),
    ("Gestió d'Or 2010 S.L.", "Valls", None, "https://www.paginasamarillas.es/search/gestorias-administrativas/all-ma/tarragona/all-is/valls/all-ba/all-pu/all-nc/1"),
    ("G&S Advocats i Economistes", "Valls", None, "https://www.gsassessors.com/"),
    ("Arils Advocats i Economistes", "Valls", "hola@arilsadvocats.com", "https://arilsassessorscom.wordpress.com/"),
    ("Gestora París S.L.", "Valls", None, "https://www.paginasamarillas.es/f/valls/gestora-paris-s-l-_009331331_000000003.html"),
    ("BS Legalcompta S.L.", "Valls", None, "https://empresite.eleconomista.es/BS-LEGALCOMPTA.html"),
    ("Cullere i Associats", "Valls", None, "https://www.gestorias.es/tarragona/valls"),
    ("Agroxarxa S.L.", "Valls", "valls@agroxarxa.com", "https://www.agroxarxa.com/contacte/"),
    ("Sugranyes Assessors (Valls)", "Valls", "valls@sugranyes.com", "https://www.sugranyes.com/contacte/"),
    ("Social Lab Assessors S.L.", "Valls", None, "https://www.social-lab.com/contacto-y-situacion/"),
    ("Parés i Aubia Assessor d'Empresa S.L.", "Valls", "info@paresiaubia.com", "https://www.cylex.es/valls/pares-i-aubia-assessor-d'empresa-s-l--12875193.html"),
    ("Assessoria Jubany S.C.P.", "Valls", None, "https://www.qdq.com/assessoria-jubany-100542"),

    ("Gestora de Gremis del Baix Penedès", "El Vendrell", "info@gestoradegremis.com", "https://www.gestoradegremis.com/"),
    ("Consulting Jayna Economistas y Auditores S.L.", "El Vendrell", None, "https://www.asesorias-empresas.es/directorio-asesorias/tarragona/el-vendrell/consulting-jayna/"),
    ("Barnadas Assessors S.L.", "El Vendrell", "info@barnadasassessors.com", "https://barnadasassessors.com/?page_id=60"),
    ("Quiron Assessors S.C.P.", "El Vendrell", None, "https://www.infoasesorias.es/quiron-assessors/"),
    ("Beltrán & Tarradellas S.L.", "El Vendrell", "soporte@beltrantarradellas.com", "https://beltrantarradellas.com/"),
    ("Norma 3 Assessors", "El Vendrell", "angels@norma3.com", "https://www.paginasamarillas.es/f/el-vendrell/norma-3-assessors_009363672_000000002.html"),
    ("Gestoria Mañé Assessors", "El Vendrell", "gestoriavendrell@gestoriamane.com", "https://www.gestoriamane.com/"),
    ("Plana Gestors S.L.", "El Vendrell", "info@planagestors.com", "https://www.gestoriaplanagestors.es/es/contacto"),
    ("Asde Assessors 2002 (El Vendrell)", "El Vendrell", "asde@asdeonline.com", "https://asdeonline.com/es/contactar/"),
    ("Joan López Assessors i Consultors S.L.", "El Vendrell", None, "https://www.paginasamarillas.es/search/asesorias-fiscales/all-ma/tarragona/all-is/el-vendrell/all-ba/all-pu/all-nc/1"),
    ("Consultoria Jurisa S.L.", "El Vendrell", None, "https://jurisa.com/"),
    ("Rovira Díaz Assessors", "El Vendrell", "info@roviradiaz.com", "https://roviradiaz.com/es/contacto/"),
    ("Assessoria Baix Penedès S.L.", "El Vendrell", None, "https://app.citvendrell.cat/empresa/assessoria-baix-peneds"),
    ("Tavi Assessors", "El Vendrell", None, "https://www.cylex.es/el-vendrell/tavi-assessors-12977502.html"),
    ("Astime", "El Vendrell", None, "https://www.sunegocio.com/pro/astime-coma-ruga-3101546/"),
    ("Asesoria Grup Vandellos", "El Vendrell", None, "https://www.gestorias.es/tarragona/el-vendrell/asesoria-grup-vandellos-16812"),
    ("DCMayer Asesores", "El Vendrell", "info@dcmayer.com", "https://www.dcmayer.com/contacto/"),
    ("Miguel Ángel Zubiria Cantarero", "El Vendrell", None, "https://www.gestorias.pro/miguel-angel-zubiria-cantarero-el-vendrell/"),
    ("Deceme Consulting S.L.", "El Vendrell", None, "https://empresite.eleconomista.es/DECEME-CONSULTING.html"),
    ("Gómez & Jané Assessors", "El Vendrell", None, "https://www.gomez-jane.com/contacto/"),

    ("Ramon Assessors", "Cambrils", "info@ramonassessors.com", "https://ramonassessors.com/es/contacto/"),
    ("Pujals & Cia Assessors", "Cambrils", "pujals@pujalsassessors.com", "https://pujalsassessors.com/contacto/"),
    ("Compta Clar Cambrils", "Cambrils", "ccc@cccambrils.cat", "https://cccambrils.cat/contacto/"),
    ("Molla Rudiez & Oliva", "Cambrils", "info@mollarudiezyoliva.com", "https://www.mollarudiezyoliva.com/contacto/despacho-en-cambrils/"),
    ("Sangenís Gestió i Serveis", "Cambrils", "fiscal@sangenisgestio.es", "https://www.gestorias.es/tarragona/cambrils/sangenis-gestio-i-serveis-2625"),
    ("Gestrams", "Cambrils", "info@gestrams.com", "https://gestrams.com/contacto/"),
    ("Perdrix-Sole Assessors", "Cambrils", "info@perdrix-sole.com", "https://perdrix-sole.com/"),
    ("Pla Comptable S.L.", "Cambrils", "placomptable@gmail.com", "https://www.placomptable.com/"),
    ("Zagin S.L.", "Cambrils", None, "https://www.paginasamarillas.es/f/cambrils/zagin-s-l_232590729_000000001.html"),
    ("Vancorblas Gestión", "Cambrils", None, "https://www.gestorias.es/tarragona/cambrils"),
    ("Reinllop", "Cambrils", None, "https://www.gestorias.es/tarragona/cambrils"),
    ("Mas Gestió Cambrils", "Cambrils", None, "https://www.gestorias.es/tarragona/cambrils/mas-gestio-cambrils-2090"),
    ("Ce Consulting Cambrils", "Cambrils", None, "https://www.gestorias.es/tarragona/cambrils/ce-consulting-cambrils-13804"),
    ("SA Consulting (Cambrils)", "Cambrils", None, "https://www.gestorias.es/tarragona/cambrils/sa-consulting-17907"),
    ("Escoda Balcells Consultors", "Cambrils", "jescoda@mugenat.es", "https://escodabalcells.es/"),
    ("Serra Assessors", "Cambrils", "agenciaserra@agenciaserra.cat", "https://www.agenciaserra.cat/es/contacto/"),
    ("G&G Gestió i Serveis", "Cambrils", "nuria@gestiogg.com", "https://www.gestiogg.cat/es/"),
    ("GR Consultors", "Cambrils", None, "https://www.holded.com/es/asesorias/cambrils"),
    ("De Donato Advocats", "Cambrils", None, "https://www.paginasamarillas.es/a/asesorias-laborales/tarragona/cambrils/"),
    ("Monur Enginyers S.L.P.", "Cambrils", None, "https://www.paginasamarillas.es/search/gestorias-administrativas/all-ma/tarragona/all-is/cambrils/all-ba/all-pu/all-nc/1"),
]


def emails_existentes():
    vistos = set()
    for path in ("cola_envios.xlsx", "contactos_catalan.xlsx", "contactos_castellano.xlsx",
                 "contactos_catalan_ronda2.xlsx", "contactos_ronda3.xlsx", "contactos_ronda4.xlsx"):
        try:
            wb = load_workbook(path)
        except FileNotFoundError:
            continue
        ws = wb.active
        h = [c.value for c in ws[1]]
        if "email" not in h:
            continue
        idx = h.index("email")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx]:
                vistos.add(str(row[idx]).strip().lower())
    return vistos


vistos = emails_existentes()
wb = load_workbook("cola_envios.xlsx")
ws = wb.active

nuevas = duplicadas = sin_email = 0
for empresa, ciudad, email, fuente in RONDA8:
    if not email:
        ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", "", "sin email", "", "", fuente])
        sin_email += 1
        continue
    if email.strip().lower() in vistos:
        duplicadas += 1
        continue
    vistos.add(email.strip().lower())
    ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", email, "", "", FECHA_HOY, fuente])
    nuevas += 1

wb.save("cola_envios.xlsx")
print(f"Ronda 8: {nuevas} nuevas con email (fecha {FECHA_HOY}), {duplicadas} duplicadas omitidas, {sin_email} sin email.")
