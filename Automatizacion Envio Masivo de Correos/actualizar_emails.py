"""Rellena email + fuente en contactos_gestorias.xlsx a partir de la búsqueda de contacto
público de cada gestoría (julio 2026). Las que no tienen email verificado quedan en blanco."""
from openpyxl import load_workbook

DATOS = {
    "Gestoria Ferrer S.L.P.": ("info@gestoriaferrer.cat", "https://gestoriaferrer.cat/"),
    "Gestoria Martí Albó": ("martialbo@martialbo.com", "https://www.gestoria-contable.es/gestoria-marti-albo_171214364-000000001/"),
    "Pere Isern Assessor S.L.": ("despatx@pereisern.cat", "https://www.pereisern.cat/es/contacto"),
    "Bufete Vidal & Sánchez, S.L.": ("info@bufetevidalsanchez.com", "https://www.bufetevidalsanchez.com/contact-with-google-map/"),
    "Gestoria Piqué S.L.": ("am@pique.cat", "http://www.pique.cat/contacto.php"),
    "Marimon Assessors": ("marimonassessors@marimonassessors.com", "https://marimonassessors.com/contacto/"),
    "Gestoria Boronat SLP": ("info@gestoriaboronat.com", "http://m.gestoriaboronat.com/contacto.html"),
    "Asesoria Gest Sabadell": ("asesorgarcia1@gmail.com", "https://gestsabadell.com/contacto/"),
    "AM Gestoria Assessoria": ("administracion@am-gestoria.es", "https://www.am-gestoria.es/"),
    "Assessoria Millan": ("asesoriamillan@apttcb.es", "https://www.paginasamarillas.es/f/sabadell/asesoria-millan_197942162_000000001.html"),
    "Gabinete Contable Vallès S.L.": ("rafaela@gcvalles.com", "https://www.gcvalles.com/contacto/"),
    "TS Gestió - Gestoría y Asesoría": ("tsgestio@tsgestio.com", "https://tsgestio.com/"),
    "Croman Asesoría Sabadell": ("croman@cromansl.com", "https://croman.es/"),
    "Serveis d'Empresa Sabadell": ("sescomptable@ses.cat", "https://www.ses.cat/contacto/"),
    "Gestoría Morillas-Azogue & Asociados SL": ("info@gestoriamorillas.com", "https://www.gestoriamorillas.com/contacto/"),
    "Assessoria Administrativa Amela SLP": ("gestoriaamela@gmail.com", "https://www.einforma.com/informacion-empresa/gestoria-administrativa-amela-slp"),
    "Gestoria Lexitus Vallès S.L.": ("lserrano@lexitusvalles.com", "https://www.lexitusvalles.net/"),
    "Gestoria Riera SL": ("grupriera@grupriera.cat", "https://www.grupriera.cat/es/"),
    "Gestoria Santaló SLP": ("info@santillanaalos.com", "http://santillanaalos.com/aviso-legal-3/"),
    "Gestoria Multi-Gest SLP": ("gamultigest@gestors.net", "https://empresite.eleconomista.es/GESTORIA-MULTI-GEST-SLP.html"),
    "Gestoria Gil Sellarès S.L.": ("gestoria@gestoriagilsellares.com", "https://www.gestoriagilsellares.com/contacto/"),
    "Gestoria Can Roca S.L.": ("info@vallparadis.com", "https://vallparadis.com/en/"),
    "Gestoria i Assessoria Vilanova SLP": ("info@gestoriavilanova.com", "https://www.gestoriavilanova.com/"),
    "Gestoria Beclan Asesores": ("fincasbeclan@gmail.com", "https://fincasbeclan.com/"),
    "Entre Trámites": ("info@entretramites.com", "https://entretramites.com/"),
    "Dc3 Asesores Legales y Gestores Tributarios": ("info@dc3asesores.com", "https://dc3asesores.com/en/index.html"),
    "Ortín & Asociados": ("ortinyasociados@ortinyasociados.com", "https://www.ortinyasociados.com/contacto/"),
    "Cross Asesores": ("crossasesores@crossasesores.com", "https://crossasesores.com/contacto/"),
    "Conat Consultores": ("info@conatconsultores.com", "https://www.conatconsultores.com/"),
    "Mir Gestoría Administrativa": ("correo@gestoriamir.com", "https://www.gestoriamir.com/"),
    "Gestoría Continente": ("continente@gestoriacontinente.es", "https://www.gestoriacontinente.es/contacto-asesoria-continente-zaragoza"),
    "M&D Gestoría": ("recepcion@mydgestoria.com", "https://mydgestoria.com/contacto/"),
    "Ignacio Ballesta Asesores": ("laboral@asesor-laboral.es", "https://www.asesor-laboral.es/"),
    "J. M. Pardo - Gestoría Administrativa": ("gestoria@jmpardo.es", "https://jmpardo.es/"),
    "Gestoría Botos": ("contacto@asesoriabotos.com", "https://gestoriabotos.com/contacto-horario-de-atencion-al-cliente/"),
    "Gestoria Galtes": ("administracio@gestoriagaltes.com", "http://www.gestoriagaltes.com/ca"),
    "Oliva & Miguel Gestió": ("info@olivamiguelgestio.com", "https://olivamiguelgestio.com/contacto/"),
    "Gestoria Bernadas": ("info@bernadas-assessoria.com", "https://www.gestoria-bernadas.com/es/contacto/"),
    "Grup Carles, Gestió i Projectes": ("gcarles@gcarles.com", "https://grupcarles.com/es/contacto"),
    "Casas Gestió": ("info@casasgestio.com", "https://www.casasgestio.com/es/contacto"),
}

wb = load_workbook("contactos_gestorias.xlsx")
ws = wb.active

headers = [c.value for c in ws[1]]
col = {name: i + 1 for i, name in enumerate(headers)}
if "fuente_email" not in col:
    nueva = ws.max_column + 1
    ws.cell(row=1, column=nueva, value="fuente_email")
    col["fuente_email"] = nueva

encontrados = 0
for row in range(2, ws.max_row + 1):
    empresa = ws.cell(row=row, column=col["empresa"]).value
    if empresa in DATOS:
        email, fuente = DATOS[empresa]
        ws.cell(row=row, column=col["email"], value=email)
        ws.cell(row=row, column=col["fuente_email"], value=fuente)
        encontrados += 1

wb.save("contactos_gestorias.xlsx")
print(f"Actualizadas {encontrados} filas con email.")
