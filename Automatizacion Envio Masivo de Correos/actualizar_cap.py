"""Reparte de nuevo fecha_prevista en cola_envios.xlsx usando el nuevo DAILY_CAP=95
(el límite real de Hostinger es 100/día; dejamos 5 de margen), añade a la cola (como
informativas, sin fecha) las gestorías de la última ronda sin email encontrado -para
que no se pierdan del histórico-, y reconstruye maestro_gestorias.xlsx reflejando
todo el cambio."""
import datetime
from openpyxl import Workbook, load_workbook

DAILY_CAP = 95
PRIMER_DIA = datetime.date(2026, 7, 14)

# (empresa, ciudad, fuente) de la última ronda (Mollet/Cornellà/Sant Boi/El Prat/Viladecans/Molins de Rei) sin email
SIN_EMAIL_NUEVAS = [
    ("Gestoría Viaplana (Viaplana Multigestió)", "Mollet del Vallès", "https://www.gestorias.es/barcelona/mollet-del-valles/gestoria-viaplana-470"),
    ("Assessoria i Serveis Empresarials Costa", "Mollet del Vallès", "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("AR Asesores Mollet", "Mollet del Vallès", "https://www.gestoriavalles.es/contacto/"),
    ("GCA Asesoría", "Mollet del Vallès", "https://www.paginasamarillas.es/f/mollet-del-valles/gca_001804467_000000001.html"),
    ("Gypesa", "Mollet del Vallès", "https://www.gestorias.es/barcelona/mollet-del-valles/gypesa-9131"),
    ("Gestoría Vallès", "Mollet del Vallès", "https://www.gestorias.es/barcelona/mollet-del-valles/gestoria-valles-17920"),
    ("Valles Gestió", "Mollet del Vallès", "https://www.gestorias.es/barcelona/mollet-del-valles/valles-gestio-16561"),
    ("Poch Assessors", "Mollet del Vallès", "https://www.paginasamarillas.es/f/mollet-del-valles/poch-assessors_020831442_000000003.html"),
    ("Mayolas Assessors d'Empreses", "Mollet del Vallès", "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Diperex S.L.", "Mollet del Vallès", "https://www.einforma.com/informacion-empresa/diperex-slp"),
    ("Bemtronic Online S.L.", "Mollet del Vallès", "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Assessoria Joan Mercadal S.L.", "Mollet del Vallès", "https://www.citiservi.es/barcelona/assessoria-joan-mercadal-mollet-del-valles__930002_67.html"),
    ("Salvador Lopez Molina", "Mollet del Vallès", "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),
    ("Rosa Bravo López", "Mollet del Vallès", "https://www.paginasamarillas.es/a/asesorias-laborales/barcelona/mollet-del-valles/"),
    ("Ecofinancial Group Consulting 86 S.L.", "Mollet del Vallès", "https://www.paginasamarillas.es/a/asesores-y-asesorias/barcelona/mollet-del-valles/"),

    ("J.M. Medina S.L.", "Cornellà de Llobregat", "https://www.gestorias.es/barcelona/cornella-de-llobregat/j-m-medina-1673"),
    ("Asesoría Roser Camps S.L.", "Cornellà de Llobregat", "https://www.paginasamarillas.es/f/cornella-de-llobregat/asesoria-roser-camps-s-l-_014410344_000000001.html"),
    ("Gestión y Asesoramientos Cornellà", "Cornellà de Llobregat", "https://www.gestorias.es/barcelona/cornella-de-llobregat/gestion-y-asesoramientos-cornella-9711"),
    ("D.M. Asesoría", "Cornellà de Llobregat", "https://firmania.es/cornella-de-llobregat/dm-asesor%C3%ADa-1693298"),
    ("Berule Consult S.L.", "Cornellà de Llobregat", "https://firmania.es/cornella-de-llobregat/berule-consult-sl-1797127"),

    ("Gestoria Abella Gestió", "Sant Boi de Llobregat", "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/gestoria-abella-gestio_000581736_000000001.html"),
    ("Gabinet Assessor Ros", "Sant Boi de Llobregat", "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/gabinet-assessor-ros_172254500_000000001.html"),
    ("Bufet Assessor ADEC S.L.P.", "Sant Boi de Llobregat", "https://infonif.economia3.com/ficha-empresa/bufet-assessor-adec-slp"),
    ("Assessoria Integral de Sant Boi S.L.", "Sant Boi de Llobregat", "https://empresite.eleconomista.es/ASSESSORIA-INTEGRAL-SANT-BOI.html"),
    ("Vikmer Assessors S.L.P.", "Sant Boi de Llobregat", "https://www.gestorias.es/barcelona/sant-boi-de-llobregat/vikmer-assessors-19018"),
    ("Bgr Assessors", "Sant Boi de Llobregat", "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/bgr-assessors_159456037_000000001.html"),
    ("Gramalla XXI Assessors al seu Servei", "Sant Boi de Llobregat", "https://www.gramallaxxi.es/?lang=es"),
    ("Aba Serveis Empresarials S.L.", "Sant Boi de Llobregat", "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/aba-serveis-empresarials-s-l-_008982217_000000001.html"),
    ("Asesoria Coope S.L.", "Sant Boi de Llobregat", "https://www.asesoriacoope.es/es/"),
    ("I.M.S. Assessors Economics i Juridics S.L.", "Sant Boi de Llobregat", "https://www.paginasamarillas.es/a/asesoria-de-empresas/barcelona/sant-boi-de-llobregat/"),
    ("ASEMRECA S.L.U.", "Sant Boi de Llobregat", "https://www.paginasamarillas.es/f/sant-boi-de-llobregat/asemreca-s-l-u-_196028658_000000002.html"),
    ("Arias Assessors (Sant Boi)", "Sant Boi de Llobregat", "https://www.carakter.org/arias-assessors"),
    ("La Asesoría Sant Boi", "Sant Boi de Llobregat", "https://laasesoriasantboi.com/contacto/"),

    ("Piera Asesorías y Servicios S.L.", "El Prat de Llobregat", "https://www.asesoriapiera.com/"),
    ("Oficina de Gestión Centro S.L.P.", "El Prat de Llobregat", "https://www.qdq.com/oficina-de-gestion-centro-514265"),
    ("Tapia Gestoría Administrativa", "El Prat de Llobregat", "https://www.infoasesorias.es/tapia-gestoria-administrativa/"),
    ("K2 Assessors", "El Prat de Llobregat", "https://directorio.guia33.com/item/k2-assessors-el-prat/"),
    ("MCC·GISE Asesoría", "El Prat de Llobregat", "https://asesoriamcc.es/"),
    ("Asesoria Gestoria Integral de Pymes S.L.", "El Prat de Llobregat", "https://www.cylex.es/el-prat-de-llobregat/asesoria-gestoria-integral-de-pymes-s-l--12503696.html"),
    ("Gestoria de la Cámara-Jonama S.L.", "El Prat de Llobregat", "https://www.gestorias.es/barcelona/el-prat-de-llobregat/gestoria-de-la-camara-jonama-3764"),
    ("Fernández Rangel S.L.", "El Prat de Llobregat", "https://www.gestorias.es/barcelona/el-prat-de-llobregat"),
    ("P&G Consultores", "El Prat de Llobregat", "http://www.pygconsultores.net/"),
    ("Asesoria Ricart S.L.", "El Prat de Llobregat", "https://empresite.eleconomista.es/ASESORIA-RICART.html"),
    ("Asesoria Teerre S.L.", "El Prat de Llobregat", "https://www.einforma.com/informacion-empresa/asesoria-teerre"),
    ("Plaça Pau Casals S.C.P.", "El Prat de Llobregat", "https://www.paginasamarillas.es/f/el-prat-de-llobregat/placa-pau-casals-s-c-p-_200511459_000000002.html"),

    ("Servicio Asesor Garantizado S.L.", "Viladecans", "https://www.paginasamarillas.es/f/viladecans/servicio-asesor-garantizado-s-l-_204687479_000000001.html"),
    ("DRI Assessoria", "Viladecans", "https://driassessoria.com/"),
    ("Asesoría Laboral Geslab S.L.", "Viladecans", "https://www.paginasamarillas.es/f/viladecans/asesoria-laboral-geslab-s-l-_021206677_000000001.html"),
    ("Baugar Gestora de Servicios", "Viladecans", "https://es.kompass.com/c/baugar-gestora-de-servicios/es1284495/"),
    ("Asesoría Areny", "Viladecans", "https://www.paginasamarillas.es/f/viladecans/asesoria-areny_008694648_000000001.html"),
    ("Fer Gestions 2009", "Viladecans", "https://empresite.eleconomista.es/FER-GESTIONS-2009.html"),
    ("Millán Gestió", "Viladecans", "https://www.infoasesorias.es/millan-gestio/"),
    ("Gómez i Carvacho Assessors", "Viladecans", "https://gomezcarvacho.com/nosotros/"),
    ("JJ & Assessors", "Viladecans", "https://assessoriajj.com/"),
    ("Arias Assessors (Viladecans)", "Viladecans", "https://www.carakter.org/arias-assessors"),
    ("Asesoría Empresarial Landa", "Viladecans", "https://aelanda.es/"),
    ("Gestram Associats 99", "Viladecans", "https://www.gestorias.es/barcelona/viladecans/gestram-associats-99-6966"),
    ("Sm Gestió", "Viladecans", "https://www.gestorias.es/barcelona/viladecans/sm-gestio-7263"),
    ("Paola Baquerizo Paladines", "Viladecans", "https://www.gestorias.es/barcelona/viladecans/2"),
    ("Rafael C. Gil", "Viladecans", "https://www.gestorias.es/barcelona/viladecans/2"),
    ("GIS, Grup Ip S.L.", "Viladecans", "https://www.gestorias.es/barcelona/viladecans/gis-grup-ip-16315"),

    ("Fàbrega Consultors", "Molins de Rei", "https://www.fabregaconsultors.com/"),
    ("Joan Tresserra Assessors S.L.", "Molins de Rei", "https://www.tresserra.cat/"),
    ("Bonafonte López S.C.P.", "Molins de Rei", "https://www.paginasamarillas.es/f/molins-de-rei/bonafonte-lopez-s-c-p-_180378630_000000001.html"),
    ("Solgemp Asesores", "Molins de Rei", "https://www.solgemp.com/"),
    ("Ferransa Asesores", "Molins de Rei", "https://ferransa.com/asesoria-fiscal-espana-molins-de-rei/"),
    ("Vernet Assessors i Associats S.L.", "Molins de Rei", "https://www.paginasamarillas.es/f/molins-de-rei/vernet-assessors-i-associats-s-l-_021411855_000000001.html"),
    ("Servei d'Assessorament Empresarial i Consultors S.L.", "Molins de Rei", "https://www.einforma.com/informacion-empresa/servei-assessorament-empresarial-consultors"),
    ("NINOAFIC S.A.", "Molins de Rei", "https://www.citiservi.es/barcelona/ninoafic-molins-de-rei__926211_940.html"),
    ("Grup Gestor Molins de Rei", "Molins de Rei", "https://www.cylex.es/molins-de-rei/grup-gestor-molins-de-rei-12661757.html"),
]

wb = load_workbook("cola_envios.xlsx")
ws = wb.active
headers = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(headers)}

empresas_en_cola = {ws.cell(row=r, column=col["empresa"]).value for r in range(2, ws.max_row + 1)}

añadidas = 0
for empresa, ciudad, fuente in SIN_EMAIL_NUEVAS:
    if empresa in empresas_en_cola:
        continue
    ws.append([f"equip de {empresa}", empresa, ciudad, "catalan", "", "sin email", "", "", fuente])
    añadidas += 1

pendientes_rows = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(row=r, column=col["estado"]).value != "enviado"
                   and ws.cell(row=r, column=col["email"]).value]

for i, r in enumerate(pendientes_rows):
    dia = PRIMER_DIA + datetime.timedelta(days=i // DAILY_CAP)
    ws.cell(row=r, column=col["fecha_prevista"], value=dia.isoformat())

wb.save("cola_envios.xlsx")
print(f"cola_envios.xlsx: +{añadidas} filas 'sin email' históricas, "
      f"{len(pendientes_rows)} filas con email reprogramadas a {DAILY_CAP}/día desde {PRIMER_DIA.isoformat()}.")

# --- reconstruir maestro_gestorias.xlsx ---
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "maestro"
ws_out.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio",
               "fecha_prevista", "fuente_email", "comentario"])


def copiar_desde(path, excluir_pendientes_con_email=False):
    wbx = load_workbook(path)
    wsx = wbx.active
    h = [c.value for c in wsx[1]]
    c = {n: i for i, n in enumerate(h)}
    for row in wsx.iter_rows(min_row=2, values_only=True):
        estado = row[c["estado"]]
        email = row[c["email"]]
        if excluir_pendientes_con_email and estado != "enviado" and email:
            continue
        estado_final = "enviado" if estado == "enviado" else ("sin email" if not email else "pendiente")
        ws_out.append([
            row[c["nombre"]], row[c["empresa"]], row[c["ciudad"]],
            "castellano" if "castellano" in path else "catalan",
            email or "", estado_final, row[c["fecha_envio"]] if estado == "enviado" else "",
            "", row[c["fuente_email"]] if "fuente_email" in c else "", "",
        ])


copiar_desde("contactos_catalan.xlsx")
copiar_desde("contactos_castellano.xlsx")
copiar_desde("contactos_catalan_ronda2.xlsx", excluir_pendientes_con_email=True)

wb_cola = load_workbook("cola_envios.xlsx")
ws_cola = wb_cola.active
h = [c.value for c in ws_cola[1]]
c = {n: i for i, n in enumerate(h)}
for row in ws_cola.iter_rows(min_row=2, values_only=True):
    if row[c["estado"]] == "enviado":
        estado_final = "enviado"
    elif not row[c["email"]]:
        estado_final = "sin email"
    else:
        estado_final = "pendiente"
    ws_out.append([
        row[c["nombre"]], row[c["empresa"]], row[c["ciudad"]], row[c["idioma"]],
        row[c["email"]] or "", estado_final, row[c["fecha_envio"]] or "", row[c["fecha_prevista"]] or "",
        row[c["fuente_email"]] or "", "",
    ])

wb_out.save("maestro_gestorias.xlsx")
print(f"maestro_gestorias.xlsx reconstruido con {ws_out.max_row - 1} filas.")
