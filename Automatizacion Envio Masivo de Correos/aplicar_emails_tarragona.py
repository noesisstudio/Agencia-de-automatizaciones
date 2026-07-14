"""Aplica los 5 emails de Tarragona que el usuario añadió al maestro y que se
perdieron en la reconstrucción anterior. Actualiza contactos_catalan_ronda2.xlsx,
los traslada a cola_envios.xlsx con fecha_prevista, y reconstruye el maestro."""
import datetime
from openpyxl import Workbook, load_workbook

DAILY_CAP = 95
PRIMER_DIA = datetime.date(2026, 7, 14)

EMAILS = {
    "Gestoria Font": "info@soldeges.com",
    "Campdepadros Gestoria": "gestoria@campdepadros.cat",
    "Corporació Raif S.L.": "info@habirent2005.com",
    "Asde Assessors 2002": "asde@asdeonline.com",
    "Vives i Roig S.L.": "info@vivesiroig.cat",
}

wb_r2 = load_workbook("contactos_catalan_ronda2.xlsx")
ws_r2 = wb_r2.active
hr2 = [c.value for c in ws_r2[1]]
cr2 = {n: i + 1 for i, n in enumerate(hr2)}

filas_nuevas_para_cola = []
for r in range(2, ws_r2.max_row + 1):
    empresa = ws_r2.cell(row=r, column=cr2["empresa"]).value
    if empresa in EMAILS and not ws_r2.cell(row=r, column=cr2["email"]).value:
        email = EMAILS[empresa]
        ws_r2.cell(row=r, column=cr2["email"], value=email)
        ciudad = ws_r2.cell(row=r, column=cr2["ciudad"]).value
        nombre = ws_r2.cell(row=r, column=cr2["nombre"]).value
        fuente = ws_r2.cell(row=r, column=cr2["fuente_email"]).value if "fuente_email" in cr2 else ""
        filas_nuevas_para_cola.append((nombre, empresa, ciudad, email, fuente))

wb_r2.save("contactos_catalan_ronda2.xlsx")
print(f"Actualizadas en ronda2: {[f[1] for f in filas_nuevas_para_cola]}")

wb = load_workbook("cola_envios.xlsx")
ws = wb.active
h = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(h)}

for nombre, empresa, ciudad, email, fuente in filas_nuevas_para_cola:
    ws.append([nombre, empresa, ciudad, "catalan", email, "", "", "", fuente])

pendientes_rows = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(row=r, column=col["estado"]).value != "enviado"
                   and ws.cell(row=r, column=col["email"]).value]

for i, r in enumerate(pendientes_rows):
    dia = PRIMER_DIA + datetime.timedelta(days=i // DAILY_CAP)
    ws.cell(row=r, column=col["fecha_prevista"], value=dia.isoformat())

wb.save("cola_envios.xlsx")

# --- reconstruir maestro_gestorias.xlsx ---
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "maestro"
ws_out.append(["nombre", "empresa", "ciudad", "idioma", "email", "estado", "fecha_envio",
               "fecha_prevista", "fuente_email", "comentario"])


def copiar_desde(path, excluir_pendientes_con_email=False):
    wbx = load_workbook(path)
    wsx = wbx.active
    hx = [c.value for c in wsx[1]]
    cx = {n: i for i, n in enumerate(hx)}
    for row in wsx.iter_rows(min_row=2, values_only=True):
        estado = row[cx["estado"]]
        email = row[cx["email"]]
        if excluir_pendientes_con_email and estado != "enviado" and email:
            continue
        estado_final = "enviado" if estado == "enviado" else ("sin email" if not email else "pendiente")
        ws_out.append([
            row[cx["nombre"]], row[cx["empresa"]], row[cx["ciudad"]],
            "castellano" if "castellano" in path else "catalan",
            email or "", estado_final, row[cx["fecha_envio"]] if estado == "enviado" else "",
            "", row[cx["fuente_email"]] if "fuente_email" in cx else "", "",
        ])


copiar_desde("contactos_catalan.xlsx")
copiar_desde("contactos_castellano.xlsx")
copiar_desde("contactos_catalan_ronda2.xlsx", excluir_pendientes_con_email=True)

wb_cola = load_workbook("cola_envios.xlsx")
ws_cola = wb_cola.active
hc = [c.value for c in ws_cola[1]]
cc = {n: i for i, n in enumerate(hc)}
for row in ws_cola.iter_rows(min_row=2, values_only=True):
    if row[cc["estado"]] == "enviado":
        estado_final = "enviado"
    elif not row[cc["email"]]:
        estado_final = "sin email"
    else:
        estado_final = "pendiente"
    ws_out.append([
        row[cc["nombre"]], row[cc["empresa"]], row[cc["ciudad"]], row[cc["idioma"]],
        row[cc["email"]] or "", estado_final, row[cc["fecha_envio"]] or "", row[cc["fecha_prevista"]] or "",
        row[cc["fuente_email"]] or "", "",
    ])

wb_out.save("maestro_gestorias.xlsx")
print(f"maestro_gestorias.xlsx reconstruido con {ws_out.max_row - 1} filas.")
