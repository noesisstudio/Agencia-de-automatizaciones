from __future__ import annotations

import csv
import io
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.core.auth import get_current_user
from app.core.security import decrypt_nif
from app.database import get_db
from app.models import Client, Invoice, InvoiceStatus, User
from app.schemas import SheetsConfigUpdate, SheetsStatusResponse
from app.services.pdf_generator import generate_invoice_pdf
from app.services.sheets_config import get_integration_settings, resolve_sheets_config
from app.services.sheets_writer import sync_invoices, test_connection

router = APIRouter(tags=["export"])

# Columnas completas que las empresas necesitan para su contabilidad.
EXPORT_HEADERS = [
    "Nº Factura",
    "Cliente",
    "NIF",
    "Fecha emisión",
    "Vencimiento",
    "Estado",
    "Base imponible",
    "Cuota IVA",
    "Retención IRPF",
    "Total",
    "Carpeta",
    "Concepto / Notas",
    "Archivo origen",
]


def _filtered_invoices(
    db: Session,
    status: str | None,
    folder_id: int | None,
    search: str | None,
    client_ids: list[int] | None = None,
    invoice_ids: list[int] | None = None,
) -> list[Invoice]:
    q = db.query(Invoice).filter(Invoice.is_deleted.is_(False))
    if invoice_ids:
        q = q.filter(Invoice.id.in_(invoice_ids))
    if status:
        try:
            q = q.filter(Invoice.status == InvoiceStatus(status))
        except ValueError as e:
            raise HTTPException(400, f"Estado no válido: {status}") from e
    if folder_id:
        q = q.filter(Invoice.folder_id == folder_id)
    if client_ids:
        q = q.filter(Invoice.client_id.in_(client_ids))
    if search and search.strip():
        term = f"%{search.strip()}%"
        q = q.filter(
            (Invoice.number.ilike(term)) | (Invoice.client.has(Client.name.ilike(term)))
        )
    return q.order_by(Invoice.created_at.desc()).all()


def _invoice_values(inv: Invoice) -> list[Any]:
    nif = decrypt_nif(inv.client.encrypted_nif) if inv.client else None
    return [
        inv.number or "",
        inv.client.name if inv.client else "",
        nif or "",
        inv.issue_date or "",
        inv.due_date or "",
        inv.status.value,
        round(float(inv.base_amount or 0), 2),
        round(float(inv.iva_amount or 0), 2),
        round(float(inv.irpf_amount or 0), 2),
        round(float(inv.total_amount or 0), 2),
        inv.folder.name if inv.folder else "",
        (inv.notes or "").replace("\n", " ").strip(),
        inv.source_filename or "",
    ]


@router.get("/export/invoices/csv")
def export_csv(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    status: str | None = Query(None),
    folder_id: int | None = Query(None),
    search: str | None = Query(None),
    client_id: Annotated[list[int] | None, Query()] = None,
    invoice_id: Annotated[list[int] | None, Query()] = None,
) -> StreamingResponse:
    rows = _filtered_invoices(db, status, folder_id, search, client_id, invoice_id)

    buf = io.StringIO()
    buf.write("\ufeff")  # BOM para que Excel reconozca UTF-8 (acentos)
    writer = csv.writer(buf, delimiter=";")  # ';' = separador estándar Excel en España
    writer.writerow(EXPORT_HEADERS)
    for inv in rows:
        writer.writerow(_invoice_values(inv))
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=facturas.csv"},
    )


@router.get("/export/invoices/excel")
def export_excel(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    status: str | None = Query(None),
    folder_id: int | None = Query(None),
    search: str | None = Query(None),
    client_id: Annotated[list[int] | None, Query()] = None,
    invoice_id: Annotated[list[int] | None, Query()] = None,
) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise HTTPException(501, "openpyxl no instalado") from e

    rows = _filtered_invoices(db, status, folder_id, search, client_id, invoice_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(EXPORT_HEADERS)
    for col, _h in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    money_cols = {7, 8, 9, 10}  # Base, IVA, IRPF, Total
    for inv in rows:
        ws.append(_invoice_values(inv))

    for r in range(2, ws.max_row + 1):
        for c in money_cols:
            ws.cell(row=r, column=c).number_format = "#,##0.00 €"

    # Anchos de columna automáticos (aproximados) y cabecera congelada.
    widths = [14, 28, 14, 14, 14, 12, 14, 12, 14, 14, 16, 36, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_HEADERS))}1"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=facturas.xlsx"},
    )


def _status_payload(db: Session) -> SheetsStatusResponse:
    cfg = resolve_sheets_config(db)
    return SheetsStatusResponse(
        configured=cfg.configured,
        credentials_detected=cfg.credentials_detected,
        has_spreadsheet_id=bool(cfg.spreadsheet_id.strip()),
        auto_sync=cfg.auto_sync,
        sheet_name=cfg.sheet_name,
        spreadsheet_id=cfg.spreadsheet_id,
        spreadsheet_url=cfg.spreadsheet_url,
        service_account_email=cfg.service_account_email,
    )


@router.get("/export/sheets/status", response_model=SheetsStatusResponse)
def sheets_status_endpoint(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
) -> SheetsStatusResponse:
    return _status_payload(db)


@router.put("/export/sheets/config", response_model=SheetsStatusResponse)
def sheets_config_endpoint(
    body: SheetsConfigUpdate,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
) -> SheetsStatusResponse:
    row = get_integration_settings(db)
    if body.google_spreadsheet_id is not None:
        row.google_spreadsheet_id = body.google_spreadsheet_id.strip()
    if body.google_sheet_name is not None:
        row.google_sheet_name = body.google_sheet_name.strip()
    if body.sheets_auto_sync is not None:
        row.sheets_auto_sync = body.sheets_auto_sync
    db.commit()
    return _status_payload(db)


@router.post("/export/sheets/test")
def sheets_test_endpoint(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
) -> dict[str, Any]:
    return test_connection(resolve_sheets_config(db))


@router.post("/export/sheets/sync")
def sheets_sync_endpoint(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    status: str | None = Query(None),
    folder_id: int | None = Query(None),
    search: str | None = Query(None),
    client_id: Annotated[list[int] | None, Query()] = None,
) -> dict[str, Any]:
    cfg = resolve_sheets_config(db)
    if not cfg.configured:
        raise HTTPException(
            400,
            "Google Sheets no está configurado. Copia el JSON de la cuenta de servicio "
            "en backend/credentials/ y define el ID de la hoja en Configuración.",
        )
    rows = _filtered_invoices(db, status, folder_id, search, client_id)
    try:
        return sync_invoices(cfg, rows)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"Error al sincronizar con Google Sheets: {e}") from e


@router.get("/invoices/{invoice_id}/pdf")
def download_pdf(
    invoice_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
) -> FileResponse:
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Factura no encontrada")
    folder = inv.folder.name if inv.folder else "General"
    path = generate_invoice_pdf(inv, folder, db=db)
    inv.pdf_path = str(path)
    db.commit()
    media = "application/pdf" if path.suffix.lower() == ".pdf" else "text/html"
    return FileResponse(path, filename=path.name, media_type=media)


@router.get("/invoices/{invoice_id}/pdf/preview")
def preview_pdf(
    invoice_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
) -> FileResponse:
    return download_pdf(invoice_id, db, _)
